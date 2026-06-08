"""
CEPS API server - Railway / cloud deployment
Lokalne: python server.py
Render: automaticky pres Procfile

POZADAVKY (requirements.txt):
    requests
    (zadne extra knihovny - pouziva jen Python stdlib)
"""
from http.server import HTTPServer, ThreadingHTTPServer, BaseHTTPRequestHandler
import json, requests, xml.etree.ElementTree as ET, os, time, io, threading
from urllib.parse import urlparse, parse_qs
from datetime import datetime, timedelta, timezone
# Pouzivame vlastni rychly XLSX streamer pres zipfile + iterparse
# (openpyxl je 30-50x pomalejsi - parsoval 120s misto 3s)

# Nacti .env soubor (pokud existuje)
def _load_env():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"): continue
                if "=" in line:
                    k, v = line.split("=", 1)
                    os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
_load_env()

BASE_URL  = "https://www.ceps.cz/_layouts/CepsData.asmx"
NAMESPACE = "https://www.ceps.cz/CepsData/"
DATA_NS   = "https://www.ceps.cz/CepsData/StructuredData/1.0"

# ENTSO-E
ENTSOE_URL   = "https://web-api.tp.entsoe.eu/api"
ENTSOE_TOKEN = os.environ.get("ENTSOE_TOKEN", "")
CZ_DOMAIN    = "10YCZ-CEPS-----N"
DE_LU_DOMAIN = "10Y1001A1001A82H"  # DE-LU bidding zone (po Oct 2018)

# Regelleistung.net (DE) aFRR ENERGY market
RL_BASE_URL = "https://www.regelleistung.net/apps/cpp-publisher/api/v2/tenders/results/anonymous"
# Fallback (older endpoint name that some installations may use):
RL_BASE_URL_ALT = "https://www.regelleistung.net/apps/crds/api/v2/tenders/results/anonymous"
RL_CACHE_TTL_SEC = 30 * 60   # 30 minut
RL_REQUEST_TIMEOUT = 60      # XLSX muze byt vetsi soubor
_RL_CACHE = {}               # {(productType, market, date): (timestamp, data)}
_RL_REFRESH_LOCK = threading.Lock()
_RL_REFRESH_INFLIGHT = set()


def _request_with_retry(func, *args, retries=3, delay=1.5, **kwargs):
    last_err = None
    for attempt in range(retries):
        try:
            return func(*args, **kwargs)
        except (requests.exceptions.SSLError,
                requests.exceptions.ConnectionError,
                requests.exceptions.Timeout) as e:
            last_err = e
            print(f"  [retry {attempt+1}/{retries}] {type(e).__name__}", flush=True)
            if attempt < retries - 1:
                time.sleep(delay * (attempt + 1))
    raise last_err

def call_ceps(method, params):
    px = "".join(f"  <{k}>{v}</{k}>\n" for k, v in params.items() if v != "")
    body = f"""<?xml version="1.0" encoding="utf-8"?>
<soap:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
  xmlns:xsd="http://www.w3.org/2001/XMLSchema"
  xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/">
  <soap:Body>
    <{method} xmlns="{NAMESPACE}">
{px}    </{method}>
  </soap:Body>
</soap:Envelope>"""
    r = _request_with_retry(
        requests.post, BASE_URL,
        data=body.encode("utf-8"),
        headers={"Content-Type": "text/xml; charset=utf-8",
                 "SOAPAction": f'"{NAMESPACE}{method}"'},
        timeout=25
    )
    return r.text, r.status_code

def parse_ceps(xml_text):
    root = ET.fromstring(xml_text)
    ns = DATA_NS
    col_map = {}
    for s in root.iter(f"{{{ns}}}serie"):
        if s.get("id") and s.get("name"):
            col_map[s.get("id")] = s.get("name")
    if not col_map:
        for s in root.iter("serie"):
            if s.get("id") and s.get("name"):
                col_map[s.get("id")] = s.get("name")
    items = list(root.iter(f"{{{ns}}}item")) or list(root.iter("item"))
    rows = []
    for item in items:
        row = {"date": item.get("date")}
        for vid, name in col_map.items():
            row[name] = item.get(vid)
        if not col_map:
            for k, v in item.attrib.items():
                if k != "date": row[k] = v
        rows.append(row)
    return {"columns": list(col_map.values()), "rows": rows}

# ============================================================
# ENTSO-E
# ============================================================

def call_entsoe(params):
    if not ENTSOE_TOKEN:
        raise RuntimeError("ENTSOE_TOKEN neni nastaveny")
    p = dict(params)
    p["securityToken"] = ENTSOE_TOKEN
    r = _request_with_retry(requests.get, ENTSOE_URL, params=p, timeout=25)
    return r.text, r.status_code

def parse_entsoe_xml(xml_text, period_start_utc):
    body = xml_text
    import re
    body_no_ns = re.sub(r'\sxmlns="[^"]+"', '', body, count=1)
    try:
        root = ET.fromstring(body_no_ns)
    except ET.ParseError:
        return []
    points = []
    for ts_node in root.findall(".//TimeSeries"):
        period = ts_node.find("Period")
        if period is None: continue
        time_interval = period.find("timeInterval")
        if time_interval is None: continue
        start_text = time_interval.find("start").text
        res_text = period.find("resolution").text
        res_min = 15
        m = re.match(r"PT(\d+)M", res_text)
        if m: res_min = int(m.group(1))
        try:
            start_dt = datetime.strptime(start_text.replace("Z", ""), "%Y-%m-%dT%H:%M")
        except Exception:
            continue
        for pt in period.findall("Point"):
            pos_el = pt.find("position")
            qty_el = pt.find("quantity")
            if pos_el is None or qty_el is None: continue
            try:
                pos = int(pos_el.text)
                qty = float(qty_el.text)
            except Exception:
                continue
            ts = start_dt + timedelta(minutes=(pos-1) * res_min)
            points.append({"ts": ts.strftime("%Y-%m-%dT%H:%MZ"), "value": qty})
    return points

def fmt_entsoe_period(d):
    return d.strftime("%Y%m%d%H%M")


# ============================================================
# Regelleistung.net (DE) aFRR ENERGY bidy
# ============================================================

def _rl_cache_get(key, allow_stale=False):
    entry = _RL_CACHE.get(key)
    if not entry: return None
    ts, data = entry
    age = time.time() - ts
    if age > RL_CACHE_TTL_SEC and not allow_stale:
        return None
    return data

def _rl_cache_age(key):
    entry = _RL_CACHE.get(key)
    if not entry: return None
    return time.time() - entry[0]

def _rl_cache_set(key, data):
    _RL_CACHE[key] = (time.time(), data)


def fetch_regelleistung_xlsx(product_type, market, delivery_date):
    """Stahne XLSX z regelleistung.net cpp-publisher API. Vraci bytes."""
    params = {
        "productType":  product_type,
        "market":       market,
        "exportFormat": "xlsx",
        "deliveryDate": delivery_date,
    }
    last_error = None
    for url in (RL_BASE_URL, RL_BASE_URL_ALT):
        try:
            r = _request_with_retry(
                requests.get, url,
                params=params, timeout=RL_REQUEST_TIMEOUT,
                headers={"User-Agent": "Mozilla/5.0 (compatible; afrr-dashboard)"}
            )
            if r.status_code != 200:
                last_error = f"HTTP {r.status_code} from {url}: {r.text[:200]}"
                continue
            if not r.content or len(r.content) < 100:
                last_error = f"Empty response from {url} ({len(r.content)} bytes)"
                continue
            print(f"  -> Regelleistung: success from {url}", flush=True)
            return r.content
        except Exception as e:
            last_error = f"Exception {url}: {e}"
            continue
    raise RuntimeError(f"Regelleistung all endpoints failed: {last_error}")


def _xlsx_iter_rows(xlsx_bytes):
    """Streamuje radky XLSX jako list stringu, BEZ openpyxl.
    Pouziva primy ZIP + iterparse - 30-50x rychlejsi nez openpyxl read_only.

    XLSX je v podstate ZIP s XML soubory. Potrebujeme:
    - xl/sharedStrings.xml: tabulka stringu (ulozena oddelene od bunek)
    - xl/worksheets/sheet1.xml: vlastni data
    """
    import zipfile
    import re as _re

    XL_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

    with zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r") as zf:
        # 1) Nacti sharedStrings.xml (pokud existuje)
        shared_strings = []
        try:
            with zf.open("xl/sharedStrings.xml") as ss_f:
                for _, elem in ET.iterparse(ss_f, events=("end",)):
                    tag = elem.tag
                    if tag == XL_NS + "si":
                        parts = []
                        for t in elem.iter(XL_NS + "t"):
                            if t.text: parts.append(t.text)
                        shared_strings.append("".join(parts))
                        elem.clear()
        except KeyError:
            shared_strings = []

        # 2) Stream sheet1.xml a vrat radky jako pole stringu
        sheet_path = None
        for name in zf.namelist():
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                sheet_path = name
                break
        if sheet_path is None:
            return

        # Helper: A1 styl letter -> 0-based column index
        def col_index(ref):
            m = _re.match(r"([A-Z]+)", ref)
            if not m: return 0
            letters = m.group(1)
            idx = 0
            for ch in letters:
                idx = idx * 26 + (ord(ch) - ord("A") + 1)
            return idx - 1

        with zf.open(sheet_path) as sh_f:
            current_row = []
            for event, elem in ET.iterparse(sh_f, events=("start", "end")):
                tag = elem.tag

                if event == "start" and tag == XL_NS + "row":
                    current_row = []

                elif event == "end" and tag == XL_NS + "c":
                    cell_ref = elem.get("r", "")
                    cell_type = elem.get("t", "")
                    v_el = elem.find(XL_NS + "v")
                    is_el = elem.find(XL_NS + "is")

                    if v_el is not None and v_el.text is not None:
                        if cell_type == "s":
                            try:
                                val = shared_strings[int(v_el.text)]
                            except (IndexError, ValueError):
                                val = ""
                        elif cell_type == "b":
                            val = "TRUE" if v_el.text == "1" else "FALSE"
                        else:
                            val = v_el.text
                    elif is_el is not None:
                        parts = []
                        for t in is_el.iter(XL_NS + "t"):
                            if t.text: parts.append(t.text)
                        val = "".join(parts)
                    else:
                        val = ""

                    col_idx = col_index(cell_ref)
                    while len(current_row) < col_idx:
                        current_row.append("")
                    current_row.append(val)
                    elem.clear()

                elif event == "end" and tag == XL_NS + "row":
                    yield current_row
                    elem.clear()
                    current_row = []


def parse_afrr_energy_xlsx(xlsx_bytes):
    """Parser RESULT_LIST_ANONYMOUS pro aFRR ENERGY market.
    Pouziva rychly streaming XML parser misto openpyxl.

    Realne sloupce v XLSX:
        DELIVERY_DATE, TYPE_OF_RESERVES, PRODUCT,
        ENERGY_PRICE_[EUR/MWh], ENERGY_PRICE_PAYMENT_DIRECTION,
        OFFERED_CAPACITY_[MW], ALLOCATED_CAPACITY_[MW], COUNTRY, NOTE

    Format produktu: 'POS_069' / 'NEG_069' (1-indexovane: POS_069 = 17:00-17:15).
    """
    import re as _re

    rows_iter = _xlsx_iter_rows(xlsx_bytes)
    try:
        headers = next(rows_iter)
    except StopIteration:
        return {"slots": {}, "directions_available": [], "raw_columns": []}

    headers = [h.strip() if h else "" for h in headers]
    headers_lower = [h.lower() for h in headers]

    def find_col(*candidates):
        for cand in candidates:
            cand_l = cand.lower()
            for i, h in enumerate(headers_lower):
                if h == cand_l: return i
        for cand in candidates:
            cand_l = cand.lower()
            for i, h in enumerate(headers_lower):
                if cand_l in h: return i
        return -1

    idx_product   = find_col("PRODUCT", "PRODUCT_NAME")
    idx_volume    = find_col("OFFERED_CAPACITY_[MW]", "OFFERED_CAPACITY",
                             "OFFERED_ENERGY_VOLUME_MW", "VOLUME_MW")
    idx_price     = find_col("ENERGY_PRICE_[EUR/MWh]", "ENERGY_PRICE",
                             "OFFERED_ENERGY_PRICE_EUR_MWH", "PRICE_EUR_MWH")
    idx_payment   = find_col("ENERGY_PRICE_PAYMENT_DIRECTION", "PAYMENT_DIRECTION")
    idx_country   = find_col("COUNTRY")
    idx_reserves  = find_col("TYPE_OF_RESERVES", "RESERVE_TYPE")
    idx_allocated = find_col("ALLOCATED_CAPACITY_[MW]", "ALLOCATED_CAPACITY")

    if idx_product < 0 or idx_volume < 0 or idx_price < 0:
        return {
            "slots": {}, "directions_available": [], "raw_columns": headers,
            "_error": (f"Pozadovane sloupce nenalezeny. "
                       f"product_idx={idx_product}, vol_idx={idx_volume}, "
                       f"price_idx={idx_price}")
        }

    raw = {}
    directions_seen = set()
    sample_products = []
    seen_products = set()
    skipped = {"non_afrr": 0, "non_de": 0, "bad_product": 0, "bad_qh": 0,
               "bad_value": 0, "bad_direction": 0}
    total_rows = 0
    max_idx = max(idx_product, idx_volume, idx_price)

    for row in rows_iter:
        total_rows += 1
        if not row or len(row) <= max_idx:
            continue
        product = (row[idx_product] or "").strip()
        volume_s = (row[idx_volume] or "").strip()
        price_s  = (row[idx_price] or "").strip()
        if not product or not volume_s or not price_s:
            continue

        # Filter aFRR
        if idx_reserves >= 0 and len(row) > idx_reserves:
            rtype = (row[idx_reserves] or "").strip().lower()
            if rtype and "afrr" not in rtype:
                skipped["non_afrr"] += 1
                continue

        # Filter DE only
        if idx_country >= 0 and len(row) > idx_country:
            ctry = (row[idx_country] or "").strip().upper()
            if ctry and ctry != "DE":
                skipped["non_de"] += 1
                continue

        if product not in seen_products and len(sample_products) < 8:
            sample_products.append(product)
            seen_products.add(product)

        # Smer
        direction = None
        pu = product.upper()
        if pu.startswith("POS"):
            direction = "POS"
        elif pu.startswith("NEG"):
            direction = "NEG"
        elif idx_payment >= 0 and len(row) > idx_payment:
            pay = (row[idx_payment] or "").strip().upper()
            if "GRID_TO_PROVIDER" in pay:
                direction = "POS"
            elif "PROVIDER_TO_GRID" in pay:
                direction = "NEG"

        if direction is None:
            skipped["bad_direction"] += 1
            continue

        # QH index
        qh_idx = None
        m = _re.match(r"(?:POS|NEG)_(\d{1,3})", product, _re.IGNORECASE)
        if not m:
            m = _re.search(r"QH[_\-]?(\d{1,3})", product, _re.IGNORECASE)
        if not m:
            m = _re.search(r"(\d{1,3})(?!.*\d)", product)
        if m:
            try:
                qh_idx = int(m.group(1))
            except ValueError:
                pass

        if qh_idx is None or qh_idx < 1 or qh_idx > 96:
            skipped["bad_qh"] += 1
            continue

        # 1-indexovani: POS_001 = 00:00-00:15, POS_069 = 17:00-17:15
        start_min = (qh_idx - 1) * 15
        sh, sm = divmod(start_min, 60)
        eh, em = divmod(start_min + 15, 60)
        if eh == 24: eh, em = 0, 0
        slot_key = f"{sh:02d}:{sm:02d}-{eh:02d}:{em:02d}"

        try:
            vol_f = float(volume_s.replace(",", "."))
            price_f = float(price_s.replace(",", "."))
        except (ValueError, TypeError):
            skipped["bad_value"] += 1
            continue

        slot_data = raw.setdefault(slot_key, {"POS": [], "NEG": []})
        slot_data[direction].append({"price": price_f, "volume_mw": vol_f})
        directions_seen.add(direction)

    # Merit-order sort + kumulativni MW pro graf
    slots_processed = {}
    for slot_key, dirs in raw.items():
        slot_out = {}
        for dir_key, bids in dirs.items():
            if not bids:
                continue
            bids.sort(key=lambda x: x["price"])
            cum = 0.0
            ladder = []
            for b in bids:
                cum += b["volume_mw"]
                ladder.append({
                    "price":     round(b["price"], 4),
                    "volume_mw": round(b["volume_mw"], 4),
                    "cum_mw":    round(cum, 4),
                })
            slot_out[dir_key.lower()] = {
                "bids":      ladder,
                "total_mw":  round(cum, 4),
                "min_price": ladder[0]["price"],
                "max_price": ladder[-1]["price"],
                "count":     len(ladder),
            }
        if slot_out:
            slots_processed[slot_key] = slot_out

    return {
        "slots":     slots_processed,
        "directions_available": sorted(directions_seen),
        "raw_columns": headers,
        "_debug": {
            "total_rows": total_rows,
            "sample_products": sample_products,
            "skipped": skipped,
            "parser": "fast-iterparse",
            "col_indices": {
                "product": idx_product, "volume": idx_volume, "price": idx_price,
                "payment": idx_payment, "country": idx_country,
                "reserves": idx_reserves, "allocated": idx_allocated,
            },
        },
    }


def _rl_refresh_in_background(delivery_date):
    """Stahne XLSX v background threadu. Idempotent."""
    key = ("aFRR", "ENERGY", delivery_date)
    with _RL_REFRESH_LOCK:
        if key in _RL_REFRESH_INFLIGHT:
            return
        _RL_REFRESH_INFLIGHT.add(key)

    def worker():
        try:
            print(f"  -> [bg] Regelleistung refresh (XLSX): {delivery_date}", flush=True)
            t0 = time.time()
            xlsx_bytes = fetch_regelleistung_xlsx("aFRR", "ENERGY", delivery_date)
            t_fetch = time.time() - t0
            parsed = parse_afrr_energy_xlsx(xlsx_bytes)
            parsed["date"] = delivery_date
            parsed["fetched_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
            _rl_cache_set(key, parsed)
            print(f"  -> [bg] OK fetched={t_fetch:.1f}s "
                  f"total={time.time()-t0:.1f}s slots={len(parsed.get('slots', {}))}",
                  flush=True)
        except Exception as e:
            print(f"  -> [bg] FAIL: {e}", flush=True)
        finally:
            with _RL_REFRESH_LOCK:
                _RL_REFRESH_INFLIGHT.discard(key)

    threading.Thread(target=worker, daemon=True).start()


def get_afrr_energy_data(delivery_date):
    """Stale-while-revalidate cache."""
    key = ("aFRR", "ENERGY", delivery_date)

    fresh = _rl_cache_get(key, allow_stale=False)
    if fresh is not None:
        out = dict(fresh); out["_cache"] = "hit"; out["_age_sec"] = int(_rl_cache_age(key) or 0)
        return out

    stale = _rl_cache_get(key, allow_stale=True)
    if stale is not None:
        age = _rl_cache_age(key) or 0
        print(f"  -> Regelleistung: cache stale (age={int(age)}s), bg refresh + vracim stale", flush=True)
        _rl_refresh_in_background(delivery_date)
        out = dict(stale); out["_cache"] = "stale"; out["_age_sec"] = int(age)
        return out

    print(f"  -> Regelleistung XLSX fetch (sync): aFRR ENERGY {delivery_date}", flush=True)
    t0 = time.time()
    xlsx_bytes = fetch_regelleistung_xlsx("aFRR", "ENERGY", delivery_date)
    t_fetch = time.time() - t0
    print(f"     downloaded {len(xlsx_bytes)} bytes in {t_fetch:.1f}s", flush=True)

    parsed = parse_afrr_energy_xlsx(xlsx_bytes)
    parsed["date"] = delivery_date
    parsed["fetched_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    print(f"     parsed in {time.time()-t0:.1f}s total: "
          f"{len(parsed.get('slots', {}))} slots, "
          f"directions={parsed.get('directions_available')}", flush=True)

    _rl_cache_set(key, parsed)
    out = dict(parsed); out["_cache"] = "miss"; out["_age_sec"] = 0
    return out


# ============================================================
# HTTP Handler
# ============================================================

class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {args[0]} {args[1]}", flush=True)

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors(); self.end_headers()

    def do_HEAD(self):
        self.send_response(200)
        self.send_header("Content-Type", "text/plain")
        self._cors(); self.end_headers()

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET,POST,OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type,X-API-Key")

    def do_POST(self):
        """Přijímá data od externích zdrojů (kamarád, partner, atd.)
        
        POST /ingest
        Header: X-API-Key: <secret>  (nebo ?key=secret v URL)
        Body: JSON s libovolnými daty
        
        Data uloží do /tmp/ingest_<timestamp>_<source>.json
        Volitelně: nastav INGEST_KEY env var pro ochranu
        Volitelně: nastav INGEST_WEBHOOK env var pro forward (Discord/Slack)
        """
        try:
            parsed = urlparse(self.path)
            qs = parse_qs(parsed.query)
            
            if parsed.path != "/ingest":
                self._json({"error": "POST only supported on /ingest"}, 404)
                return
            
            # API Key check (header or query param)
            expected_key = os.environ.get("INGEST_KEY", "")
            if expected_key:
                provided = self.headers.get("X-API-Key", "") or qs.get("key", [""])[0]
                if provided != expected_key:
                    self._json({"error": "invalid API key"}, 401)
                    return
            
            # Read body
            content_length = int(self.headers.get("Content-Length", 0))
            if content_length > 50 * 1024 * 1024:  # 50 MB max
                self._json({"error": "body too large (max 50MB)"}, 413)
                return
            
            body = self.rfile.read(content_length) if content_length > 0 else b"{}"
            content_type = self.headers.get("Content-Type", "").lower()
            
            # Parse podle Content-Type
            source = qs.get("source", ["unknown"])[0]
            data_type = qs.get("type", ["data"])[0]
            
            parsed_data = None
            raw_text = None
            
            if "application/json" in content_type:
                try:
                    parsed_data = json.loads(body.decode("utf-8"))
                except json.JSONDecodeError as e:
                    self._json({"error": f"invalid JSON: {e}"}, 400)
                    return
            elif "text/csv" in content_type or "text/plain" in content_type:
                raw_text = body.decode("utf-8", errors="replace")
            else:
                # Binary (XLSX, PDF, atd.) - uloží raw
                pass
            
            # Save to disk + memory store
            ts = datetime.now().strftime("%Y%m%d-%H%M%S")
            safe_source = "".join(c for c in source if c.isalnum() or c in "._-")[:50]
            
            # Memory store (last 100 messages per source)
            if "_INGEST_STORE" not in globals():
                globals()["_INGEST_STORE"] = {}
            store = globals()["_INGEST_STORE"]
            if safe_source not in store:
                store[safe_source] = []
            
            record = {
                "ts": ts,
                "source": safe_source,
                "type": data_type,
                "content_type": content_type,
                "size": len(body),
                "data": parsed_data if parsed_data is not None else (raw_text[:1000] if raw_text else "<binary>"),
            }
            store[safe_source].append(record)
            if len(store[safe_source]) > 100:
                store[safe_source] = store[safe_source][-100:]  # keep last 100
            
            # File save (optional - /tmp survives only restart)
            try:
                file_ext = ".json" if parsed_data else (".csv" if raw_text else ".bin")
                fpath = f"/tmp/ingest_{ts}_{safe_source}_{data_type}{file_ext}"
                if parsed_data is not None:
                    with open(fpath, "w") as f:
                        json.dump(parsed_data, f, ensure_ascii=False, indent=2)
                elif raw_text is not None:
                    with open(fpath, "w") as f:
                        f.write(raw_text)
                else:
                    with open(fpath, "wb") as f:
                        f.write(body)
                print(f"[INGEST] saved {len(body)} bytes from '{safe_source}' to {fpath}", flush=True)
            except Exception as e:
                print(f"[INGEST] file save error: {e}", flush=True)
            
            # Optional webhook forward (Discord/Slack)
            webhook_url = os.environ.get("INGEST_WEBHOOK", "")
            if webhook_url:
                try:
                    summary = f"📥 Ingest from '{safe_source}' ({data_type}): {len(body)} bytes"
                    if parsed_data and isinstance(parsed_data, dict):
                        keys = list(parsed_data.keys())[:5]
                        summary += f"\nKeys: {', '.join(keys)}"
                    requests.post(webhook_url, json={"content": summary}, timeout=5)
                except Exception as e:
                    print(f"[INGEST] webhook fail: {e}", flush=True)
            
            self._json({
                "status": "ok",
                "ts": ts,
                "source": safe_source,
                "type": data_type,
                "size": len(body),
                "parsed": parsed_data is not None
            })
        except (BrokenPipeError, ConnectionResetError):
            pass
        except Exception as e:
            print(f"!!! UNCAUGHT do_POST ERROR: {self.path} - {e}", flush=True)
            import traceback
            traceback.print_exc()
            try:
                self._json({"error": "internal error", "detail": str(e)}, 500)
            except Exception:
                pass

    def do_GET(self):
        try:
            self._do_GET_inner()
        except (BrokenPipeError, ConnectionResetError):
            pass  # Klient zavrel connection - normalni
        except Exception as e:
            print(f"!!! UNCAUGHT do_GET ERROR: {self.path} - {e}", flush=True)
            import traceback
            traceback.print_exc()
            try:
                self._json({"error": "internal error", "detail": str(e)}, 500)
            except Exception:
                pass

    def _do_GET_inner(self):
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)
        g = lambda k, d="": qs.get(k, [d])[0]

        if parsed.path == "/health":
            self._json({"status": "ok", "time": datetime.now().isoformat(),
                        "version": "v42-vdt-range"}); return
        
        # Ingest read endpoints
        if parsed.path == "/ingest/list":
            store = globals().get("_INGEST_STORE", {})
            summary = {}
            for src, records in store.items():
                summary[src] = {
                    "count": len(records),
                    "latest_ts": records[-1]["ts"] if records else None,
                    "types": list(set(r["type"] for r in records))
                }
            self._json({"sources": summary, "total_sources": len(store)})
            return
        
        if parsed.path == "/ingest/data":
            source = qs.get("source", [None])[0]
            data_type = qs.get("type", [None])[0]
            limit = int(qs.get("limit", ["10"])[0])
            store = globals().get("_INGEST_STORE", {})
            if source and source in store:
                records = store[source]
                if data_type:
                    records = [r for r in records if r["type"] == data_type]
                self._json({"source": source, "records": records[-limit:]})
            else:
                self._json({"error": "source not found", "available": list(store.keys())}, 404)
            return
        
        # VDT orderbook - merged snapshot s cache
        if parsed.path == "/vdt/orderbook":
            store = globals().get("_INGEST_STORE", {})
            cache = globals().get("_VDT_OB_CACHE", {})
            records = store.get("ote-com-bridge", [])
            records = [r for r in records if r.get("type") == "vdt_orderbook"]
            
            # Cache key = počet records + last ts
            latest_ts = records[-1].get("ts") if records else None
            cache_key = (len(records), latest_ts)
            
            if cache.get("key") == cache_key and "result" in cache:
                # Cached - vrať okamžitě
                self._json(cache["result"])
                return
            
            # Merge: pro každý kontrakt najdi nejnovější hodnoty
            merged = {}
            for rec in records:
                data = rec.get("data")
                if not isinstance(data, dict): continue
                rows = data.get("rows", [])
                for row in rows:
                    k = row.get("kontrakt", {})
                    raw = k.get("raw") if isinstance(k, dict) else None
                    if not raw: continue
                    merged[raw] = row
            
            def sort_key(r):
                k = r.get("kontrakt", {})
                f = k.get("from", "99:99")
                t = k.get("type", "Z")
                return (f, t)
            sorted_rows = sorted(merged.values(), key=sort_key)
            
            result = {
                "rows": sorted_rows,
                "latest_ts": latest_ts,
                "total_kontraktu": len(sorted_rows),
                "ingest_records": len(records),
            }
            # Cache
            globals()["_VDT_OB_CACHE"] = {"key": cache_key, "result": result}
            self._json(result)
            return
        
        # Route /vdt.html, /analyza.html, /ema.html atd - generic static file serve
        if parsed.path.endswith(".html") and parsed.path != "/" and "/" not in parsed.path[1:]:
            try:
                base_dir = os.path.dirname(os.path.abspath(__file__))
                fname = parsed.path.lstrip("/")
                fpath = os.path.join(base_dir, fname)
                if os.path.exists(fpath):
                    with open(fpath, "rb") as f:
                        content = f.read()
                    self.send_response(200)
                    self.send_header("Content-Type", "text/html; charset=utf-8")
                    self.send_header("Content-Length", str(len(content)))
                    self.end_headers()
                    self.wfile.write(content)
                    return
            except Exception as e:
                print(f"static html error: {e}", flush=True)

        if parsed.path in ("/", "/index.html", "/hory.html"):
            # NEW landing: hory.html (cista verze bez Systemove soustavy)
            try:
                base_dir = os.path.dirname(os.path.abspath(__file__))
                fpath = os.path.join(base_dir, "hory.html")
                if os.path.exists(fpath):
                    with open(fpath, "rb") as f:
                        content = f.read()
                    self.send_response(200)
                    self.send_header("Content-Type", "text/html; charset=utf-8")
                    self.send_header("Content-Length", str(len(content)))
                    self.send_header("Cache-Control", "no-cache")
                    self.end_headers()
                    self.wfile.write(content)
                    return
            except Exception:
                pass
            # fallback: live_odchylky.html
            self._html(); return

        if parsed.path == "/entsoe/solar":
            self._entsoe_solar(qs); return
        
        # Actual generation per type/country (solar+wind) z ENTSO-E
        if parsed.path == "/entsoe/actual":
            try:
                country = qs.get("country", ["CZ"])[0].upper()
                gen_type = qs.get("type", ["solar"])[0].lower()
                day_str = qs.get("day", [None])[0]
                
                EIC = {
                    "CZ": "10YCZ-CEPS-----N",
                    "DE": "10Y1001A1001A83F",  # DE total
                    "AT": "10YAT-APG------L",
                    "HU": "10YHU-MAVIR----U",
                    "NL": "10YNL----------L",
                    "PL": "10YPL-AREA-----S",
                    "SK": "10YSK-SEPS-----K",
                }.get(country)
                if not EIC:
                    self._json({"error": f"unknown country {country}", "available": ["CZ","DE","AT","HU","NL"]}, 400); return
                
                psrType = {"solar": "B16", "wind": "B19", "wind_onshore": "B19", "wind_offshore": "B18"}.get(gen_type, "B16")
                
                if day_str:
                    day = datetime.strptime(day_str, "%Y-%m-%d")
                else:
                    day = datetime.now(timezone.utc).replace(tzinfo=None)
                ps = day.replace(hour=0, minute=0, second=0, microsecond=0)
                pe = ps + timedelta(hours=23)
                
                # A75 = Actual generation per type
                params = {
                    "documentType": "A75",
                    "processType": "A16",  # Realised
                    "psrType": psrType,
                    "in_Domain": EIC,
                    "periodStart": fmt_entsoe_period(ps),
                    "periodEnd": fmt_entsoe_period(pe),
                }
                
                # Cache 10 min
                cache_key = f"actual|{country}|{gen_type}|{ps.strftime('%Y-%m-%d')}"
                if "_ENTSOE_ACT_CACHE" not in globals(): globals()["_ENTSOE_ACT_CACHE"] = {}
                cache = globals()["_ENTSOE_ACT_CACHE"]
                now = time.time()
                if cache_key in cache and (now - cache[cache_key]["ts"]) < 600:
                    out = dict(cache[cache_key]["data"]); out["_cache"] = "hit"
                    self._json(out); return
                
                xml, st = call_entsoe(params)
                if st != 200:
                    self._json({"error": f"ENTSO-E status {st}", "preview": xml[:200] if isinstance(xml, str) else str(xml)[:200]}, 200); return
                
                points = parse_entsoe_xml(xml, ps)
                # Convert to {ts, value} format - same as MAGMA
                result_points = []
                for p in points:
                    result_points.append({"ts": p.get("ts") or p.get("timestamp"), "value": p.get("value")})
                
                out = {
                    "country": country,
                    "type": gen_type,
                    "day": ps.strftime("%Y-%m-%d"),
                    "points": result_points,
                    "n": len(result_points),
                }
                cache[cache_key] = {"ts": now, "data": out}
                self._json(out)
            except Exception as e:
                self._json({"error": str(e)[:200]}, 500)
            return

        if parsed.path == "/entsoe/residual-load" or parsed.path == "/entsoe/residual":
            self._entsoe_residual_load(qs); return
        
        if parsed.path == "/entsoe/afrr-cz":
            self._entsoe_afrr_cz(qs); return
        
        if parsed.path == "/smard/residual-load":
            self._smard_residual_load(qs); return

        if parsed.path == "/ote/spot":
            self._ote_spot(qs); return

        if parsed.path == "/ote/qh":
            self._ote_qh(qs); return

        if parsed.path == "/ote/dt15":
            self._ote_dt15(qs); return

        if parsed.path == "/ote/zo":
            self._ote_zo(qs); return

        if parsed.path == "/ote/yearly-profile":
            self._ote_yearly_profile(qs); return

        if parsed.path == "/ote/last7d-stats":
            self._ote_last7d_stats(qs); return

        if parsed.path == "/ote/debug":
            # Debug endpoint - ukaze stav historie a cache
            history = globals().get("_OTE_HISTORY", {})
            cache = globals().get("_OTE_SPOT_CACHE", {})
            history_summary = {}
            for date_str, hours in history.items():
                history_summary[date_str] = {
                    "count": len(hours),
                    "first": hours[0] if hours else None,
                    "last": hours[-1] if hours else None,
                }
            self._json({
                "history_dates": list(history.keys()),
                "history_summary": history_summary,
                "cache_age_sec": int(time.time() - cache.get("ts", 0)) if cache.get("ts") else None,
                "cache_has_data": cache.get("data") is not None,
            }); return

        if parsed.path == "/weather":
            self._weather(qs); return

        if parsed.path == "/metdesk/probe":
            # Probe ruzne MetDesk power endpointy
            try:
                import urllib.request, urllib.error
                api_key = os.environ.get("METDESK_API_KEY", "").strip("\n\r\t ")
                if not api_key:
                    self._json({"error": "no key"}); return
                raw_token = api_key.split(" ", 1)[1] if " " in api_key else api_key
                
                # Test ruzne kombinace location/element pro vsechny tri modely
                now_utc = datetime.now(timezone.utc)
                start = now_utc.strftime("%Y-%m-%dT00:00:00Z")
                end = (now_utc + timedelta(hours=24)).strftime("%Y-%m-%dT00:00:00Z")
                
                # Najit latest issue
                def get_latest_issue(model):
                    try:
                        req = urllib.request.Request(
                            f"https://api.metdesk.com/get/metdesk/powergen/v2/issues?model={model}",
                            headers={"Authorization": f"jwt {raw_token}"})
                        with urllib.request.urlopen(req, timeout=5) as r:
                            d = json.loads(r.read())
                        return d.get("data", [])[-1] if d.get("data") else None
                    except: return None
                
                magma_issue = get_latest_issue("magma")
                icon_issue = get_latest_issue("icon")
                
                urls_to_test = []
                # Možné endpointy pro observations/actuals
                urls_to_test.append("https://api.metdesk.com/get/metdesk/powergen/v2/observations")
                urls_to_test.append("https://api.metdesk.com/get/metdesk/powergen/v2/observations?model=magma&location=DE&location_type=country&element=solar")
                urls_to_test.append("https://api.metdesk.com/get/metdesk/powergen/v2/actuals")
                urls_to_test.append("https://api.metdesk.com/get/metdesk/powergen/v2/realised")
                urls_to_test.append("https://api.metdesk.com/get/metdesk/powergen/v2/measurements")
                urls_to_test.append("https://api.metdesk.com/get/metdesk/powergen/v2/history")
                urls_to_test.append("https://api.metdesk.com/get/metdesk/powergen/v2/historical")
                urls_to_test.append("https://api.metdesk.com/get/metdesk/powergen/v2/nowcast")
                urls_to_test.append("https://api.metdesk.com/get/metdesk/powergen/v2/analysis")
                urls_to_test.append("https://api.metdesk.com/get/metdesk/powergenobs/v2/issues")
                urls_to_test.append("https://api.metdesk.com/get/metdesk/observations/v2/issues")
                urls_to_test.append("https://api.metdesk.com/get/metdesk/powergen_actual/v2/issues")
                urls_to_test.append("https://api.metdesk.com/get/metdesk/powergen/v2/data")
                urls_to_test.append("https://api.metdesk.com/get/metdesk/powergen/v2/timeseries")
                urls_to_test.append("https://api.metdesk.com/get/metdesk/powergen/v2/")
                
                # Také zkusit starší issue (před 48h) - to bude historický forecast = "actual" po tom času
                old_issue = (now_utc - timedelta(hours=48)).strftime("%Y-%m-%dT00:00:00Z")
                urls_to_test.append(f"https://api.metdesk.com/get/metdesk/powergen/v2/forecasts?model=magma&issue={old_issue}&location=DE&location_type=country&element=solar&interval=hires&start_dtg={(now_utc - timedelta(hours=48)).strftime('%Y-%m-%dT00:00:00Z')}&end_dtg={now_utc.strftime('%Y-%m-%dT00:00:00Z')}")
                
                results = []
                for url in urls_to_test:
                    # Vytvor zkraceny label: jen relevantni parametry
                    if "?" in url:
                        params = url.split("?", 1)[1]
                        # Vytahnout jen klicove: model, location, element
                        import urllib.parse as _up
                        qp = dict(_up.parse_qsl(params))
                        short = f"{qp.get('model','?')} | {qp.get('location','-')} | {qp.get('element','-')}"
                    else:
                        short = url.split("/metdesk/")[1] if "/metdesk/" in url else url
                    try:
                        req = urllib.request.Request(url, headers={"Authorization": f"jwt {raw_token}"})
                        with urllib.request.urlopen(req, timeout=6) as r:
                            body = r.read(400).decode("utf-8", "replace")
                            # Kontrolovat data_count
                            data_count = "?"
                            try:
                                j = json.loads(body + "}" if not body.endswith("}") else body[:body.rfind("}")+1])
                                data_count = j.get("request", {}).get("data_count", "?")
                            except: pass
                            results.append({
                                "id": short,
                                "status": r.status,
                                "n": data_count,
                                "preview": body[:120]
                            })
                    except urllib.error.HTTPError as e:
                        err = ""
                        try: err = e.read().decode("utf-8")[:150]
                        except: pass
                        results.append({
                            "id": short,
                            "status": e.code,
                            "err": err
                        })
                    except Exception as e:
                        results.append({"id": short, "err": str(e)[:80]})
                
                # Filter to relevant (200 OK + 401/403 jsou zajímavé)
                interesting = [r for r in results if r.get("status", 0) in (200, 401, 403)]
                self._json({"interesting": interesting, "total_tested": len(results)})
            except Exception as e:
                self._json({"error": str(e)}, 500)
            return

        if parsed.path == "/netztransparenz/activated":
            self._ntp_activated(qs); return

        if parsed.path == "/eupowerprices/forecast":
            self._eupowerprices_forecast(qs); return

        if parsed.path == "/eupowerprices/areas":
            self._eupowerprices_areas(); return

        if parsed.path == "/metdesk/magma":
            self._metdesk_magma(qs); return

        if parsed.path == "/metdesk/weather":
            self._metdesk_weather(qs); return

        if parsed.path == "/metdesk/debug":
            # Debug: ukaz co je v env var (bez prozrazeni klice)
            k = os.environ.get("METDESK_API_KEY", "")
            self._json({
                "exists": bool(k),
                "length": len(k),
                "starts_with": k[:4] if len(k) >= 4 else "",
                "ends_with": k[-4:] if len(k) >= 4 else "",
                "has_leading_space": k.startswith(" ") if k else False,
                "has_trailing_space": k.endswith(" ") if k else False,
                "has_inner_space": " " in k.strip() if k else False,
                "num_spaces": k.count(" "),
                "has_newline": "\n" in k,
                "has_tab": "\t" in k,
                "char_codes_first5": [ord(c) for c in k[:5]],
                "char_codes_last5": [ord(c) for c in k[-5:]] if len(k) >= 5 else []
            }); return

        if parsed.path == "/wind-de":
            self._wind_de(qs); return

        if parsed.path == "/forecast/de":
            self._forecast_de(qs); return

        if parsed.path == "/regelleistung/afrr-energy":
            self._regelleistung_afrr_energy(qs); return

        if parsed.path == "/regelleistung/debug":
            self._regelleistung_debug(qs); return

        # Spot ceny (DA) z energy-charts.info pro CZ/DE
        if parsed.path == "/spot/prices":
            self._spot_prices(qs); return
        
        # OTE VDT (vnitrodenni trh, last cena per QH)
        if parsed.path == "/ote/vdt":
            self._ote_vdt(qs); return
        
        # OTE VDT range - VDT data za poslednich N dni (pro hruska.html)
        if parsed.path == "/ote/vdt/range":
            self._ote_vdt_range(qs); return
        
        # ČEPS TXT download (reálná data) - pro ema.html backtest
        if parsed.path == "/ceps_txt":
            self._ceps_txt(qs); return
        
        # Staticke HTML soubory (hruska.html, kapacity.html, live_odchylky.html)
        if parsed.path in ("/hruska.html", "/kapacity.html", "/live_odchylky.html", "/ema.html", "/odhad.html", "/eisi.html", "/ceny.html", "/fanda.html", "/api_test.html", "/spread.html"):
            try:
                fname = parsed.path.lstrip("/")
                # Hleda soubor vedle server.py
                base_dir = os.path.dirname(os.path.abspath(__file__))
                fpath = os.path.join(base_dir, fname)
                if os.path.exists(fpath):
                    with open(fpath, "rb") as f:
                        content = f.read()
                    self.send_response(200)
                    self.send_header("Content-Type", "text/html; charset=utf-8")
                    self.send_header("Content-Length", str(len(content)))
                    self.send_header("Cache-Control", "no-cache")
                    self.end_headers()
                    self.wfile.write(content)
                    return
                else:
                    self._json({"error": f"file not found: {fname}"}, 404); return
            except Exception as e:
                self._json({"error": str(e)}, 500); return

        if parsed.path != "/api":
            self._json({"error": "use /api"}, 404); return

        method = g("method", "AktualniSystemovaOdchylkaCR")
        df     = g("dateFrom")
        dt_    = g("dateTo")
        agr    = g("agregation", "MI")
        fn     = g("function",   "AVG")
        ver    = g("version",    "RT")
        para1  = g("para1", "")

        if not df:
            now = datetime.now()
            df  = now.strftime("%Y-%m-%dT00:00:00")
            dt_ = now.strftime("%Y-%m-%dT%H:%M:%S")

        params = {"dateFrom": df, "dateTo": dt_}

        if method == "AktualniSystemovaOdchylkaCR":
            params.update({"agregation": agr, "function": fn})
        elif method == "AktivaceSVRvCR":
            params.update({"aggregation": agr, "function": fn})
            if para1: params["para1"] = para1
        elif method == "AktualniCenaRE":
            # Aktualni cena regulacni energie (aFRR, mFRR+, mFRR-)
            # agregation: MI (minuta) / QH / HR
            params.update({"agregation": agr, "function": fn})
        elif method in ["Load","Generation","GenerationRES","CrossborderPowerFlows"]:
            params.update({"agregation": agr, "function": fn, "version": ver})
            if para1: params["para1"] = para1
        elif method == "ExportImportSVR":
            params.update({"agregation": agr, "function": fn})
            if para1: params["para1"] = para1

        # === CACHE pro ochranu pred banem (CEPS API rate limit) ===
        # Live data (last 4h range) -> 30s cache
        # Historicka data (dany den fixne) -> 5min cache
        cache_key = f"{method}|{df}|{dt_}|{agr}|{fn}|{ver}|{para1}"
        if not hasattr(self.__class__, '_api_cache'):
            self.__class__._api_cache = {}
        cache = self.__class__._api_cache
        
        # Pro request s end-date danes (live) cache 30s
        # Pro historicka data cache 5 min
        try:
            from datetime import datetime as _dt
            is_today = dt_.startswith(_dt.now().strftime("%Y-%m-%d"))
        except Exception:
            is_today = False
        ttl = 30 if is_today else 300
        
        now_ts = time.time()
        if cache_key in cache:
            ts, cached_data = cache[cache_key]
            if now_ts - ts < ttl:
                print(f"  -> {method}: CACHED ({int(now_ts - ts)}s old)", flush=True)
                self._json(cached_data); return
        
        try:
            xml_text, status = call_ceps(method, params)
        except Exception as e:
            print(f"  -> {method} REQUEST FAIL: {e}", flush=True)
            # Pri chybe vrat stary cache pokud existuje
            if cache_key in cache:
                print(f"  -> {method}: ČEPS fail, vracim STALE cache", flush=True)
                self._json(cache[cache_key][1]); return
            self._json({"error": f"CEPS request failed: {e}"}, 502); return

        if status != 200:
            # CEPS 500 (rate limit) -> vrat stary cache pokud existuje
            if cache_key in cache:
                print(f"  -> {method}: ČEPS {status}, vracim STALE cache", flush=True)
                self._json(cache[cache_key][1]); return
            try:
                root = ET.fromstring(xml_text)
                fs = root.find(".//{http://schemas.xmlsoap.org/soap/envelope/}faultstring")
                msg = fs.text if fs is not None else xml_text[:300]
            except Exception:
                msg = xml_text[:300]
            self._json({"error": f"CEPS {status}: {msg}"}, 502); return

        data = parse_ceps(xml_text)
        data["fetched_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        # Ulozit do cache
        cache[cache_key] = (now_ts, data)
        # Vycisti stary cache (>1h)
        if len(cache) > 200:
            stale = [k for k,(t,_) in cache.items() if now_ts - t > 3600]
            for k in stale: del cache[k]
        
        print(f"  -> {method}: {len(data['rows'])} radku, cols={data['columns']}", flush=True)
        self._json(data)

    def _entsoe_solar(self, qs):
        try:
            day_str = qs.get("day", [None])[0]
            if day_str:
                day = datetime.strptime(day_str, "%Y-%m-%d")
            else:
                day = datetime.now(timezone.utc).replace(tzinfo=None)
            ps = day.replace(hour=0, minute=0, second=0, microsecond=0)
            pe = ps + timedelta(hours=23)

            base = {
                "documentType": "A69",
                "psrType":      "B16",
                "in_Domain":    CZ_DOMAIN,
                "periodStart":  fmt_entsoe_period(ps),
                "periodEnd":    fmt_entsoe_period(pe),
            }

            da_xml, da_st = call_entsoe({**base, "processType": "A01"})
            da_points = parse_entsoe_xml(da_xml, ps) if da_st == 200 else []
            print(f"  -> ENTSO-E Solar DA: status={da_st}, points={len(da_points)}", flush=True)
            if da_st != 200:
                tok_len = len(ENTSOE_TOKEN) if ENTSOE_TOKEN else 0
                tok_preview = (ENTSOE_TOKEN[:6] + "..." + ENTSOE_TOKEN[-4:]) if tok_len > 12 else "(empty)"
                print(f"     token_len={tok_len}, token={tok_preview}", flush=True)
                print(f"     response[:300]: {da_xml[:300]}", flush=True)

            id_xml, id_st = call_entsoe({**base, "processType": "A40"})
            id_points = parse_entsoe_xml(id_xml, ps) if id_st == 200 else []
            print(f"  -> ENTSO-E Solar ID: status={id_st}, points={len(id_points)}", flush=True)

            self._json({
                "day": ps.strftime("%Y-%m-%d"),
                "day_ahead": da_points,
                "intraday":  id_points,
            })
        except Exception as e:
            print(f"  -> ENTSO-E Solar ERROR: {e}", flush=True)
            self._json({"error": str(e)}, 502)

    def _entsoe_afrr_cz(self, qs):
        """Vraci aFRR PICASSO ceny pro CZ.
        
        ENTSO-E documentType=A88 (Imbalance volume) / A85 (Imbalance prices)
        Granularita: 4 sek (ale ENTSO-E publikuje typicky 15min agregat)
        
        Query: ?date=YYYY-MM-DD (default dnes)
        Vraci: {date, qh: [{interval, hour, minute, price_pos, price_neg, vol_pos, vol_neg}, ...]}
        
        Cache 5 min.
        """
        try:
            date_str = qs.get("date", [None])[0]
            if not date_str:
                now_utc = datetime.now(timezone.utc)
                month = now_utc.month
                berlin_offset = 2 if 4 <= month <= 10 else 1
                berlin_now = now_utc + timedelta(hours=berlin_offset)
                date_str = berlin_now.strftime("%Y-%m-%d")
            
            try:
                target_date = datetime.strptime(date_str, "%Y-%m-%d")
            except ValueError:
                self._json({"error": f"Invalid date: {date_str}"}, 400); return
            
            nocache = qs.get("nocache", ["0"])[0] in ("1", "true", "yes")
            
            # Cache
            cache_key = f"_AFRR_CZ_CACHE_{date_str}"
            if cache_key not in globals():
                globals()[cache_key] = {"ts": 0, "data": None}
            cache = globals()[cache_key]
            now_ts = time.time()
            if not nocache and cache["data"] and (now_ts - cache["ts"]) < 300:
                out = dict(cache["data"]); out["_cache"] = "hit"
                self._json(out); return
            
            # ENTSO-E period - cely den (UTC) - posun pro Berlin TZ
            month = target_date.month
            berlin_offset = 2 if 4 <= month <= 10 else 1
            
            day_start_local = target_date.replace(hour=0, minute=0, second=0)
            day_end_local = day_start_local + timedelta(hours=24)
            # Konverze na UTC pro ENTSO-E (které chce UTC)
            ps = day_start_local - timedelta(hours=berlin_offset)
            pe = day_end_local - timedelta(hours=berlin_offset)
            
            period_start = fmt_entsoe_period(ps)
            period_end = fmt_entsoe_period(pe)
            
            # ENTSO-E documentType: PICASSO aFRR Imbalance Prices
            # CZ je v PICASSO od 1.6.2022
            # 
            # documentType options pro balancing:
            #   A85 - Imbalance settlement (CENA)  
            #   A86 - Imbalance volume
            #   A89 - Cross-border marginal price (XBMP)
            #
            # Pro národní imbalance prices: 
            #   documentType=A85, controlArea_Domain=10YCZ-CEPS-----N
            
            params_price = {
                "documentType": "A85",
                "controlArea_Domain": CZ_DOMAIN,
                "periodStart": period_start,
                "periodEnd": period_end,
            }
            
            print(f"  -> /entsoe/afrr-cz: fetching A85 for {date_str}", flush=True)
            try:
                price_xml, price_st = call_entsoe(params_price)
            except Exception as e:
                print(f"  -> /entsoe/afrr-cz A85 fetch FAIL: {e}", flush=True)
                self._json({"error": f"ENTSO-E A85 fetch failed: {e}"}, 502); return
            
            if price_st != 200:
                # Try alternate documentType
                print(f"  -> A85 returned {price_st}, trying A89 (XBMP)...", flush=True)
                params_xbmp = {
                    "documentType": "A89",
                    "controlArea_Domain": CZ_DOMAIN,
                    "periodStart": period_start,
                    "periodEnd": period_end,
                }
                try:
                    price_xml, price_st = call_entsoe(params_xbmp)
                except Exception as e:
                    self._json({"error": f"ENTSO-E A85+A89 both failed: {e}"}, 502); return
                
                if price_st != 200:
                    self._json({
                        "error": f"ENTSO-E both A85 and A89 returned non-200",
                        "status_A85": price_st,
                        "response_preview": price_xml[:500] if price_xml else None
                    }, 502); return
            
            # Parse XML response
            import re as _re
            import xml.etree.ElementTree as ET
            body_no_ns = _re.sub(r'\sxmlns="[^"]+"', '', price_xml, count=1)
            try:
                root = ET.fromstring(body_no_ns)
            except ET.ParseError as e:
                self._json({"error": f"XML parse: {e}", "preview": price_xml[:500]}, 502); return
            
            # Parse TimeSeries
            qh_data = []  # list of {ts, value, direction}
            
            for ts_node in root.findall(".//TimeSeries"):
                # Smer (POS / NEG / flow)
                direction = None
                flow_dir = ts_node.find("flowDirection.direction")
                if flow_dir is not None and flow_dir.text:
                    direction = flow_dir.text  # "A01"=up, "A02"=down
                else:
                    # try businessType
                    bt = ts_node.find("businessType")
                    if bt is not None: direction = bt.text
                
                # Period
                period = ts_node.find("Period")
                if period is None: continue
                ti = period.find("timeInterval")
                if ti is None: continue
                start_el = ti.find("start")
                if start_el is None: continue
                
                try:
                    period_start_dt = datetime.strptime(start_el.text.replace("Z",""), "%Y-%m-%dT%H:%M")
                except ValueError:
                    continue
                
                res_el = period.find("resolution")
                res_min = 15  # default
                if res_el is not None:
                    m = _re.match(r"PT(\d+)([MS])", res_el.text)
                    if m:
                        v, unit = int(m.group(1)), m.group(2)
                        if unit == "S": res_min = v / 60.0
                        else: res_min = v
                
                for pt in period.findall("Point"):
                    pos_el = pt.find("position")
                    price_el = pt.find("price.amount")
                    qty_el = pt.find("quantity")
                    
                    if pos_el is None: continue
                    try:
                        pos = int(pos_el.text)
                    except: continue
                    
                    value = None
                    if price_el is not None and price_el.text:
                        try: value = float(price_el.text)
                        except: pass
                    elif qty_el is not None and qty_el.text:
                        try: value = float(qty_el.text)
                        except: pass
                    
                    if value is None: continue
                    
                    ts_dt = period_start_dt + timedelta(minutes=(pos-1) * res_min)
                    # Konverze na Berlin TZ
                    ts_berlin = ts_dt + timedelta(hours=berlin_offset)
                    
                    qh_data.append({
                        "ts": ts_dt.strftime("%Y-%m-%dT%H:%MZ"),
                        "berlin_time": ts_berlin.strftime("%Y-%m-%dT%H:%M"),
                        "hour": ts_berlin.hour,
                        "minute": ts_berlin.minute,
                        "value": value,
                        "direction": direction,
                    })
            
            # Group by QH and direction
            from collections import defaultdict
            qh_grouped = defaultdict(lambda: {"price_pos": None, "price_neg": None, "vol_pos": None, "vol_neg": None})
            for d in qh_data:
                qh_key = (d["hour"], (d["minute"] // 15) * 15)
                # A01 = up/POS, A02 = down/NEG
                if d["direction"] == "A01":
                    qh_grouped[qh_key]["price_pos"] = d["value"]
                elif d["direction"] == "A02":
                    qh_grouped[qh_key]["price_neg"] = d["value"]
            
            qh_list = []
            for (h, m), vals in sorted(qh_grouped.items()):
                qh_list.append({
                    "hour": h,
                    "minute": m,
                    "interval": f"{h:02d}:{m:02d}-{(h + (1 if m==45 else 0)):02d}:{(m+15)%60:02d}",
                    "price_pos_eur": vals["price_pos"],
                    "price_neg_eur": vals["price_neg"],
                })
            
            out = {
                "date": date_str,
                "source": "ENTSO-E Transparency Platform A85",
                "qh": qh_list,
                "count": len(qh_list),
                "raw_points": len(qh_data),
                "fetched_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                "_cache": "miss",
            }
            cache["ts"] = now_ts
            cache["data"] = {k: v for k, v in out.items() if k != "_cache"}
            print(f"  -> /entsoe/afrr-cz: {len(qh_list)} QH, {len(qh_data)} raw points", flush=True)
            self._json(out)
        except Exception as e:
            print(f"  -> /entsoe/afrr-cz ERROR: {e}", flush=True)
            import traceback
            traceback.print_exc()
            self._json({"error": str(e)}, 502)

    def _entsoe_residual_load(self, qs):
        """Residual Load Forecast pro DE-LU = Demand - Wind - Solar.
        Klicovy indikator pro ceny zitra: vysoka residual load -> drahe peaky,
        nizka/zaporna -> levne nebo zaporne ceny.
        Cache 1h. Query ?days_back=30 -> historie + zítra.
        """
        try:
            days_back = int(qs.get("days_back", ["0"])[0])
            days_back = max(0, min(30, days_back))
            
            # Cache key zahrnuje days_back aby ruzne pozadavky se nepretlucily
            cache_key = f"_RESIDUAL_LOAD_CACHE_{days_back}"
            if cache_key not in globals():
                globals()[cache_key] = {"ts": 0, "data": None}
            cache = globals()[cache_key]
            now_ts = time.time()
            if cache["data"] and (now_ts - cache["ts"]) < 3600:
                out = dict(cache["data"]); out["_cache"] = "hit"
                out["_age_sec"] = int(now_ts - cache["ts"])
                self._json(out); return

            # Berlin TZ - posledních N dní zpět + zítra
            now_utc = datetime.now(timezone.utc)
            month = now_utc.month
            berlin_offset = 2 if 4 <= month <= 10 else 1
            berlin_now = now_utc + timedelta(hours=berlin_offset)
            # Start: dnešní 00:00 mínus days_back
            ps_local = berlin_now.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=days_back)
            # End: dnešní 00:00 + 48h (=zítřek+1)
            pe_local = berlin_now.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(hours=48)
            # Convert back to UTC for ENTSO-E
            ps = ps_local - timedelta(hours=berlin_offset)
            pe = pe_local - timedelta(hours=berlin_offset)

            period_start = fmt_entsoe_period(ps)
            period_end = fmt_entsoe_period(pe)

            # 1) Demand forecast (A65 = Load forecast, processType A01 = Day-Ahead)
            demand_xml, demand_st = call_entsoe({
                "documentType": "A65",
                "processType":  "A01",
                "outBiddingZone_Domain": DE_LU_DOMAIN,
                "periodStart":  period_start,
                "periodEnd":    period_end,
            })
            demand_pts = parse_entsoe_xml(demand_xml, ps) if demand_st == 200 else []
            print(f"  -> ENTSO-E Demand DE: status={demand_st}, points={len(demand_pts)}", flush=True)

            # 2) Wind onshore forecast (A69 + B19)
            wind_on_xml, wind_on_st = call_entsoe({
                "documentType": "A69",
                "processType":  "A01",
                "psrType":      "B19",  # Wind Onshore
                "in_Domain":    DE_LU_DOMAIN,
                "periodStart":  period_start,
                "periodEnd":    period_end,
            })
            wind_on_pts = parse_entsoe_xml(wind_on_xml, ps) if wind_on_st == 200 else []
            print(f"  -> ENTSO-E Wind onshore DE: status={wind_on_st}, points={len(wind_on_pts)}", flush=True)

            # 3) Wind offshore forecast (A69 + B18)
            wind_off_xml, wind_off_st = call_entsoe({
                "documentType": "A69",
                "processType":  "A01",
                "psrType":      "B18",  # Wind Offshore
                "in_Domain":    DE_LU_DOMAIN,
                "periodStart":  period_start,
                "periodEnd":    period_end,
            })
            wind_off_pts = parse_entsoe_xml(wind_off_xml, ps) if wind_off_st == 200 else []
            print(f"  -> ENTSO-E Wind offshore DE: status={wind_off_st}, points={len(wind_off_pts)}", flush=True)

            # 4) Solar forecast (A69 + B16)
            solar_xml, solar_st = call_entsoe({
                "documentType": "A69",
                "processType":  "A01",
                "psrType":      "B16",  # Solar
                "in_Domain":    DE_LU_DOMAIN,
                "periodStart":  period_start,
                "periodEnd":    period_end,
            })
            solar_pts = parse_entsoe_xml(solar_xml, ps) if solar_st == 200 else []
            print(f"  -> ENTSO-E Solar DE: status={solar_st}, points={len(solar_pts)}", flush=True)

            # Sjednoc na hodinove buckety - klic = ts string
            def to_dict(points):
                return {p["ts"]: p["value"] for p in points}

            demand_d = to_dict(demand_pts)
            wind_on_d = to_dict(wind_on_pts)
            wind_off_d = to_dict(wind_off_pts)
            solar_d = to_dict(solar_pts)

            # Vsechny ts klicove napric ctyrmi datasety
            all_ts = sorted(set(demand_d.keys()) | set(wind_on_d.keys())
                          | set(wind_off_d.keys()) | set(solar_d.keys()))

            result = []
            for ts in all_ts:
                d = demand_d.get(ts)
                wo = wind_on_d.get(ts, 0)
                woff = wind_off_d.get(ts, 0)
                s = solar_d.get(ts, 0)
                wind_total = (wo or 0) + (woff or 0)
                if d is None:
                    continue  # bez demand nemuzeme spocitat residual
                # Pokud chybi wind I solar forecast (oba = 0), preskoc
                # (forecast pro D+2 ENTSO-E nepublikuje, dostali bychom jen demand jako residual)
                if wind_total == 0 and (s or 0) == 0:
                    continue
                residual = d - wind_total - (s or 0)
                # Convert UTC ts to Berlin local time pro frontend
                dt_utc = datetime.strptime(ts.replace("Z", ""), "%Y-%m-%dT%H:%M").replace(tzinfo=timezone.utc)
                dt_berlin = dt_utc + timedelta(hours=berlin_offset)
                result.append({
                    "ts": ts,
                    "berlin_time": dt_berlin.strftime("%Y-%m-%dT%H:%M"),
                    "demand_mw": round(d, 1),
                    "wind_mw": round(wind_total, 1),
                    "solar_mw": round(s or 0, 1),
                    "residual_load_mw": round(residual, 1),
                })

            out = {
                "data": result,
                "country": "DE-LU",
                "period": {
                    "start": ps_local.strftime("%Y-%m-%dT%H:%M"),
                    "end": pe_local.strftime("%Y-%m-%dT%H:%M"),
                },
                "source": "transparency.entsoe.eu",
                "fetched_at": now_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
                "_cache": "miss",
            }
            cache["ts"] = now_ts
            cache["data"] = {k: v for k, v in out.items() if k != "_cache"}
            print(f"  -> residual-load: {len(result)} bodu", flush=True)
            self._json(out)
        except Exception as e:
            print(f"  -> /entsoe/residual-load ERROR: {e}", flush=True)
            import traceback
            traceback.print_exc()
            self._json({"error": str(e)}, 502)

    # ============================================================
    # SMARD - DE residual load (rychlejsi nez ENTSO-E, ~30 min delay)
    # ============================================================
    _SMARD_CACHE = {"ts": 0, "data": None}
    
    def _smard_residual_load(self, qs):
        """SMARD DE residual load - 15min granularita, ~30min delay.
        Stahuje z https://www.smard.de/app/chart_data/4359/DE/index_quarterhour.json
        Cache 15 min.
        """
        try:
            now_ts = time.time()
            if self._SMARD_CACHE["data"] and (now_ts - self._SMARD_CACHE["ts"]) < 900:
                out = dict(self._SMARD_CACHE["data"])
                out["_cache"] = "hit"
                self._json(out); return
            
            # 1. Stahnout index s casovymi razitky
            idx_url = "https://www.smard.de/app/chart_data/4359/DE/index_quarterhour.json"
            r_idx = requests.get(idx_url, timeout=15)
            if r_idx.status_code != 200:
                self._json({"error": f"smard index status {r_idx.status_code}", "data": []}, 502); return
            
            idx = r_idx.json()
            timestamps = idx.get("timestamps", [])
            if not timestamps:
                self._json({"error": "no smard timestamps", "data": []}, 502); return
            
            # 2. Stahnout poslednich 5 tydnu (5 souboru)
            from datetime import timezone as tz
            all_points = []
            recent_ts = sorted(timestamps)[-5:]  # poslednich 5 tydenu
            for week_ts in recent_ts:
                data_url = f"https://www.smard.de/app/chart_data/4359/DE/4359_DE_quarterhour_{week_ts}.json"
                try:
                    r_data = requests.get(data_url, timeout=15)
                    if r_data.status_code != 200: continue
                    j = r_data.json()
                    series = j.get("series", [])
                    for pt in series:
                        if not isinstance(pt, list) or len(pt) < 2: continue
                        ts_ms, val = pt[0], pt[1]
                        if val is None: continue
                        dt = datetime.fromtimestamp(ts_ms / 1000, tz=tz.utc)
                        all_points.append({
                            "ts": int(ts_ms),
                            "berlin_time": dt.strftime("%Y-%m-%dT%H:%M:%SZ"),
                            "residual_load_actual_mw": float(val),
                        })
                except Exception as e:
                    print(f"  -> SMARD week {week_ts} ERROR: {e}", flush=True)
                    continue
            
            # Setrid podle ts
            all_points.sort(key=lambda p: p["ts"])
            
            out = {
                "data": all_points,
                "source": "smard.de",
                "fetched_at": datetime.now(tz.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                "_cache": "miss",
            }
            self._SMARD_CACHE = {"ts": now_ts, "data": out}
            print(f"  -> SMARD residual-load: {len(all_points)} bodu", flush=True)
            self._json(out)
        except Exception as e:
            print(f"  -> /smard/residual-load ERROR: {e}", flush=True)
            import traceback
            traceback.print_exc()
            self._json({"error": str(e), "data": []}, 502)

    def _ote_yearly_profile(self, qs):
        """Vraci hodinovy profil (0-23) cen elektriny v CR za posledni rok.
        Pro kazdou hodinu vraci: prumer, min, max, percentily (p25, p75).
        Zdroj: api.energy-charts.info (Fraunhofer ISE), bzn=CZ.
        Cache 24h (data se meni jen 1x denne pri publikaci day-ahead).
        """
        try:
            nocache = qs.get("nocache", ["0"])[0] in ("1", "true", "yes")
            if "_OTE_YEARLY_CACHE" not in globals():
                globals()["_OTE_YEARLY_CACHE"] = {"ts": 0, "data": None}
            cache = globals()["_OTE_YEARLY_CACHE"]
            now_ts = time.time()
            # Cache 24h (86400s)
            if not nocache and cache["data"] and (now_ts - cache["ts"]) < 86400:
                out = dict(cache["data"]); out["_cache"] = "hit"
                self._json(out); return

            # Stahnu data za posledni rok z Fraunhofer Energy-Charts
            now_utc = datetime.now(timezone.utc)
            end_date = now_utc.strftime("%Y-%m-%d")
            start_date = (now_utc - timedelta(days=365)).strftime("%Y-%m-%d")
            url = f"https://api.energy-charts.info/price?bzn=CZ&start={start_date}&end={end_date}"

            print(f"  -> /ote/yearly-profile fetching {url}", flush=True)
            r = _request_with_retry(
                requests.get,
                url,
                timeout=60,
                headers={"User-Agent": "Mozilla/5.0 (compatible; ceps-dashboard)"}
            )
            if r.status_code != 200:
                self._json({"error": f"Energy-charts HTTP {r.status_code}"}, 502); return

            data = r.json()
            timestamps = data.get("unix_seconds", [])
            prices = data.get("price", [])
            if not timestamps or not prices or len(timestamps) != len(prices):
                self._json({"error": "Energy-charts: empty or invalid data"}, 502); return

            # Berlin TZ - cena za hodinu HH ve dni
            # Pro kazdy timestamp vypocitam Berlin hour a pridam do bucket
            buckets = {h: [] for h in range(24)}

            for ts, price in zip(timestamps, prices):
                if price is None:
                    continue
                # Konvertuj UTC na Berlin
                dt_utc = datetime.fromtimestamp(ts, tz=timezone.utc)
                month = dt_utc.month
                # Berlin offset: leto +2, zima +1 (bez DST presnosti)
                berlin_offset = 2 if 4 <= month <= 10 else 1
                dt_berlin = dt_utc + timedelta(hours=berlin_offset)
                hour = dt_berlin.hour
                buckets[hour].append(float(price))

            # Vypocitaj statistiky pro kazdou hodinu
            def percentile(values, p):
                if not values:
                    return None
                s = sorted(values)
                k = (len(s) - 1) * (p / 100)
                f = int(k)
                c = min(f + 1, len(s) - 1)
                if f == c:
                    return s[f]
                return s[f] + (s[c] - s[f]) * (k - f)

            profile = []
            for h in range(24):
                vals = buckets[h]
                if not vals:
                    profile.append({
                        "hour": h, "count": 0,
                        "avg": None, "min": None, "max": None,
                        "p25": None, "median": None, "p75": None
                    })
                    continue
                profile.append({
                    "hour": h,
                    "count": len(vals),
                    "avg": round(sum(vals) / len(vals), 2),
                    "min": round(min(vals), 2),
                    "max": round(max(vals), 2),
                    "p25": round(percentile(vals, 25), 2),
                    "median": round(percentile(vals, 50), 2),
                    "p75": round(percentile(vals, 75), 2),
                })

            # Globalni prumer roku (pro kontext)
            all_prices = [p for vals in buckets.values() for p in vals]
            year_avg = round(sum(all_prices) / len(all_prices), 2) if all_prices else None
            year_min = round(min(all_prices), 2) if all_prices else None
            year_max = round(max(all_prices), 2) if all_prices else None

            out = {
                "profile": profile,
                "period": {"start": start_date, "end": end_date},
                "stats": {
                    "year_avg_eur": year_avg,
                    "year_min_eur": year_min,
                    "year_max_eur": year_max,
                    "total_hours": len(all_prices),
                    "negative_hours": sum(1 for p in all_prices if p < 0),
                },
                "source": "api.energy-charts.info (Fraunhofer ISE) bzn=CZ",
                "fetched_at": now_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
                "_cache": "miss",
            }
            cache["ts"] = now_ts
            cache["data"] = {k: v for k, v in out.items() if k != "_cache"}
            print(f"  -> yearly-profile: {len(all_prices)} hodin, prumer {year_avg} EUR", flush=True)
            self._json(out)
        except Exception as e:
            print(f"  -> /ote/yearly-profile ERROR: {e}", flush=True)
            import traceback
            traceback.print_exc()
            self._json({"error": str(e)}, 502)

    def _ote_last7d_stats(self, qs):
        """Vraci statistiky CR ceny za poslednich 7 dni:
        - prumerna cena (EUR/MWh)
        - prumerny denni spread (max-min v ramci dne, prumer pres dny)
        Zdroj: api.energy-charts.info (Fraunhofer ISE), bzn=CZ.
        Cache 1 hodina.
        """
        try:
            if "_LAST7D_CACHE" not in globals():
                globals()["_LAST7D_CACHE"] = {"ts": 0, "data": None}
            cache = globals()["_LAST7D_CACHE"]
            now_ts = time.time()
            # 1 hodina cache
            if cache["data"] and (now_ts - cache["ts"]) < 3600:
                out = dict(cache["data"]); out["_cache"] = "hit"
                out["_age_sec"] = int(now_ts - cache["ts"])
                self._json(out); return

            now_utc = datetime.now(timezone.utc)
            end_date = now_utc.strftime("%Y-%m-%d")
            start_date = (now_utc - timedelta(days=7)).strftime("%Y-%m-%d")
            url = f"https://api.energy-charts.info/price?bzn=CZ&start={start_date}&end={end_date}"

            print(f"  -> /ote/last7d-stats fetching {url}", flush=True)
            try:
                r = _request_with_retry(
                    requests.get, url, timeout=30,
                    headers={"User-Agent": "Mozilla/5.0 (compatible; ceps-dashboard)"}
                )
                fetch_failed = (r.status_code != 200)
            except Exception:
                fetch_failed = True
                r = None

            if fetch_failed:
                # Stale cache fallback
                if cache["data"]:
                    out = dict(cache["data"])
                    out["_cache"] = "stale"
                    out["_age_sec"] = int(now_ts - cache["ts"])
                    self._json(out); return
                self._json({"error": "Energy-charts API failed"}, 502); return

            data = r.json()
            timestamps = data.get("unix_seconds", [])
            prices = data.get("price", [])
            if not timestamps or not prices or len(timestamps) != len(prices):
                self._json({"error": "Energy-charts: empty data"}, 502); return

            # Seskup ceny po dnech (Berlin TZ) - jen HODINOVE ceny (filtruj :00)
            # Energy-Charts vraci 15min data, ale pro DA arbitraz pouzivame hodinove
            from collections import defaultdict
            days = defaultdict(list)
            for ts, price in zip(timestamps, prices):
                if price is None:
                    continue
                dt_utc = datetime.fromtimestamp(ts, tz=timezone.utc)
                month = dt_utc.month
                berlin_offset = 2 if 4 <= month <= 10 else 1
                dt_berlin = dt_utc + timedelta(hours=berlin_offset)
                # Filtruj jen :00 minuty (1 cena per hour)
                if dt_berlin.minute != 0:
                    continue
                day_key = dt_berlin.strftime("%Y-%m-%d")
                days[day_key].append(float(price))

            # Spocitej statistiky
            all_prices = []
            day_spreads = []
            day_avgs = []
            day_mins = []
            day_maxs = []
            for day_key in sorted(days.keys()):
                day_prices = days[day_key]
                if len(day_prices) < 12:  # Potreba aspon 12h pro spread (jinak neuplny den)
                    continue
                d_min = min(day_prices)
                d_max = max(day_prices)
                d_avg = sum(day_prices) / len(day_prices)
                day_spreads.append(d_max - d_min)
                day_avgs.append(d_avg)
                day_mins.append(d_min)
                day_maxs.append(d_max)
                all_prices.extend(day_prices)

            if not all_prices:
                self._json({"error": "Energy-charts: nedostatek dat"}, 502); return

            avg_price = sum(all_prices) / len(all_prices)
            avg_spread = sum(day_spreads) / len(day_spreads) if day_spreads else 0
            avg_min = sum(day_mins) / len(day_mins) if day_mins else 0
            avg_max = sum(day_maxs) / len(day_maxs) if day_maxs else 0
            min_spread = min(day_spreads) if day_spreads else 0
            max_spread = max(day_spreads) if day_spreads else 0

            # Median spread - robustnejsi vuci outlierum
            sorted_spreads = sorted(day_spreads)
            median_spread = sorted_spreads[len(sorted_spreads)//2] if sorted_spreads else 0

            # Per-day breakdown pro debug
            per_day = []
            for day_key in sorted(days.keys()):
                dp = days[day_key]
                if len(dp) < 12:
                    continue
                per_day.append({
                    "day": day_key,
                    "hours": len(dp),
                    "min": round(min(dp), 2),
                    "max": round(max(dp), 2),
                    "spread": round(max(dp) - min(dp), 2),
                    "avg": round(sum(dp)/len(dp), 2),
                })

            out = {
                "avg_price_eur": round(avg_price, 2),
                "avg_spread_eur": round(avg_spread, 2),
                "median_spread_eur": round(median_spread, 2),
                "min_spread_eur": round(min_spread, 2),
                "max_spread_eur": round(max_spread, 2),
                "avg_min_eur": round(avg_min, 2),
                "avg_max_eur": round(avg_max, 2),
                "days_count": len(day_spreads),
                "total_hours": len(all_prices),
                "per_day": per_day,
                "period": {"start": start_date, "end": end_date},
                "source": "api.energy-charts.info (Fraunhofer ISE) bzn=CZ",
                "fetched_at": now_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
                "_cache": "miss",
            }
            cache["ts"] = now_ts
            cache["data"] = {k: v for k, v in out.items() if k != "_cache"}
            print(f"  -> last7d-stats: avg {avg_price:.1f}€, spread {avg_spread:.1f}€ "
                  f"({len(day_spreads)} dni)", flush=True)
            self._json(out)
        except Exception as e:
            print(f"  -> /ote/last7d-stats ERROR: {e}", flush=True)
            import traceback
            traceback.print_exc()
            self._json({"error": str(e)}, 502)

    def _ote_dt15(self, qs):
        """Vraci 15-min ceny z OTE-CR.cz oficialni denni trh.
        Primary: XLSX, Fallback: HTML stranky.
        Cache 5 min.
        """
        try:
            day_offset = int(qs.get("day", ["0"])[0])
            nocache = qs.get("nocache", ["0"])[0] in ("1", "true", "yes")
            
            cache_key = f"_OTE_DT15_CACHE_{day_offset}"
            if cache_key not in globals():
                globals()[cache_key] = {"ts": 0, "data": None}
            cache = globals()[cache_key]
            now = time.time()
            if not nocache and cache["data"] and (now - cache["ts"]) < 300:
                out = dict(cache["data"]); out["_cache"] = "hit"
                self._json(out); return
            
            now_utc = datetime.now(timezone.utc)
            month_now = now_utc.month
            berlin_offset = 2 if 4 <= month_now <= 10 else 1
            berlin_now = now_utc + timedelta(hours=berlin_offset)
            target_date = berlin_now + timedelta(days=day_offset)
            yyyy = target_date.year
            mm = target_date.month
            dd = target_date.day
            date_str_iso = target_date.strftime("%Y-%m-%d")
            
            unique_rows = []
            source_str = ""
            error_msg = ""
            
            # === PRIMARY: XLSX ===
            try:
                xlsx_url = (
                    f"https://www.ote-cr.cz/pubweb/attachments/01/{yyyy}/"
                    f"month{mm:02d}/day{dd:02d}/DT_15MIN_{dd:02d}_{mm:02d}_{yyyy}_CZ.xlsx"
                )
                r = _request_with_retry(
                    requests.get, xlsx_url, timeout=20,
                    headers={"User-Agent": "Mozilla/5.0 (compatible; ceps-dashboard)"}
                )
                if r.status_code == 200:
                    from openpyxl import load_workbook
                    from io import BytesIO
                    import re as _re
                    wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
                    ws = wb.active
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        if row is None or len(row) < 2: continue
                        for c_idx in range(len(row) - 1):
                            interval_val = row[c_idx]
                            if interval_val is None: continue
                            interval_str = str(interval_val).strip()
                            m = _re.match(r'^(\d{2}):(\d{2})-(\d{2}):(\d{2})$', interval_str)
                            if not m: continue
                            price_val = row[c_idx + 1]
                            if price_val is None: continue
                            try:
                                if isinstance(price_val, (int, float)):
                                    price_eur = float(price_val)
                                else:
                                    ps = str(price_val).strip().replace(' ', '').replace(',', '.').replace('\xa0', '')
                                    price_eur = float(ps)
                            except (ValueError, TypeError):
                                continue
                            rows.append({
                                "hour": int(m.group(1)),
                                "minute": int(m.group(2)),
                                "interval": f"{m.group(1)}:{m.group(2)}-{m.group(3)}:{m.group(4)}",
                                "price_eur": round(price_eur, 2),
                            })
                            break  # 1 interval per row
                    wb.close()
                    
                    seen = set()
                    for row in rows:
                        key = (row["hour"], row["minute"])
                        if key not in seen:
                            seen.add(key)
                            unique_rows.append(row)
                    unique_rows.sort(key=lambda r: (r["hour"], r["minute"]))
                    if unique_rows:
                        source_str = "ote-cr.cz XLSX"
                else:
                    error_msg = f"XLSX HTTP {r.status_code}"
            except Exception as e:
                error_msg = f"XLSX parse: {e}"
            
            # === FALLBACK: HTML parsing ===
            if not unique_rows:
                try:
                    date_url = target_date.strftime("%d.%m.%Y")
                    html_url = f"https://www.ote-cr.cz/cs/kratkodobe-trhy/elektrina/denni-trh?date={date_url}"
                    r2 = _request_with_retry(
                        requests.get, html_url, timeout=15,
                        headers={"User-Agent": "Mozilla/5.0 (compatible; ceps-dashboard)"}
                    )
                    if r2.status_code == 200:
                        import re as _re
                        html = r2.text
                        # OTE má 2 tabulky: 1) BASE/PEAK shrnutí 2) QH detail
                        # Najdi POSLEDNÍ tabulku se 4-cifernymi intervaly
                        # Pattern: jen rows s intervalem co skutečně začíná hodinami 00-23
                        # Format z OTE: <td>HH:MM-HH:MM</td><td>cena</td>...
                        # Cena ma format "162,46" nebo "1 038,450" (s mezerou)
                        # 15min cena je vzdy 2 sloupec po intervalu
                        
                        # Najdi <tbody> nebo prosto vsechny <tr> s intervalem
                        # Robustni pattern - bere prvni cislo po intervalu
                        row_pattern = _re.compile(
                            r'<td[^>]*>\s*(\d{2}):(\d{2})-(\d{2}):(\d{2})\s*</td>\s*'
                            r'<td[^>]*>\s*([-\d][\d\s,.\xa0]*?)\s*</td>',
                            _re.DOTALL
                        )
                        for m in row_pattern.finditer(html):
                            h_start = int(m.group(1))
                            min_start = int(m.group(2))
                            # 15min cena - jednoduche formatovani (mala cislo bez tisicovych mezer)
                            # Vetsi cisla (mnozstvi) maji format "1 038,450" - tj 4+ digity
                            price_raw = m.group(5).strip()
                            price_str = price_raw.replace('\xa0', '').replace(' ', '').replace(',', '.')
                            try:
                                price_eur = float(price_str)
                            except ValueError:
                                continue
                            # Cena RE 15min v CZ je typicky -200 az 500 EUR
                            # Mnozstvi je 800-1500 MWh - pokud > 1000, mozna chytame spatny sloupec
                            # ALE cena muze legitimne byt 500+. Hard limit 600.
                            if abs(price_eur) > 600:
                                continue
                            unique_rows.append({
                                "hour": h_start,
                                "minute": min_start,
                                "interval": f"{h_start:02d}:{min_start:02d}-{m.group(3)}:{m.group(4)}",
                                "price_eur": round(price_eur, 2),
                            })
                        # Dedupe + sort
                        seen = set()
                        deduped = []
                        for row in unique_rows:
                            key = (row["hour"], row["minute"])
                            if key not in seen:
                                seen.add(key)
                                deduped.append(row)
                        deduped.sort(key=lambda r: (r["hour"], r["minute"]))
                        unique_rows = deduped
                        if unique_rows:
                            source_str = "ote-cr.cz HTML fallback"
                except Exception as e:
                    error_msg += f" | HTML: {e}"
            
            if not unique_rows:
                self._json({"error": f"OTE DT15: no data - {error_msg}"}, 502); return
            
            out = {
                "date": date_str_iso,
                "day_offset": day_offset,
                "qh": unique_rows,
                "count": len(unique_rows),
                "source": source_str,
                "fetched_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                "_cache": "miss",
            }
            cache["ts"] = now
            cache["data"] = {k: v for k, v in out.items() if k != "_cache"}
            self._json(out)
        except Exception as e:
            import traceback
            self._json({"error": f"OTE DT15: {e}", "trace": traceback.format_exc()[:500]}, 500)

    def _ote_zo(self, qs):
        """OTE Výsledky zúčtování odchylek (finální cena).
        URL: /attachments/05_09_12/{year}/month{mm}/day{dd}/RPVZ_15MIN_DD_MM_YYYY_CZ.xlsx
        Query: ?day=N (offset) or ?date=YYYY-MM-DD
        Vraci: {date, qh: [{hour, minute, interval, price_eur}], count, source}
        Cache 30 min (zúčtování je finální, nemění se).
        """
        try:
            # Determine date
            date_param = qs.get("date", [None])[0]
            if date_param:
                target_date = datetime.strptime(date_param, "%Y-%m-%d")
            else:
                day_offset = int(qs.get("day", ["-1"])[0])
                now_utc = datetime.now(timezone.utc)
                month_now = now_utc.month
                berlin_offset = 2 if 4 <= month_now <= 10 else 1
                berlin_now = now_utc + timedelta(hours=berlin_offset)
                target_date = berlin_now + timedelta(days=day_offset)
            
            yyyy = target_date.year
            mm = target_date.month
            dd = target_date.day
            date_str_iso = target_date.strftime("%Y-%m-%d")
            nocache = qs.get("nocache", ["0"])[0] in ("1", "true", "yes")
            
            # Cache (30 min, zúčtování je finální)
            # ALE: skipni cache pokud má prázdné qh (mohla být chyba parsingu)
            cache_key = f"_OTE_ZO_CACHE_{date_str_iso}"
            if cache_key not in globals():
                globals()[cache_key] = {"ts": 0, "data": None}
            cache = globals()[cache_key]
            now_ts = time.time()
            if not nocache and cache["data"] and (now_ts - cache["ts"]) < 1800:
                cached_qh = cache["data"].get("qh", [])
                if len(cached_qh) > 0:  # JEN pokud cache má data
                    out = dict(cache["data"]); out["_cache"] = "hit"
                    self._json(out); return
            
            # URL varianty - název souboru
            # Hlavní: Odchylky_DD_MM_YYYY_V0_CZ.xlsx
            # V0 = první publikace, V1+ = opravy/závěrečná verze
            file_names = [
                f"Odchylky_{dd:02d}_{mm:02d}_{yyyy}_V0_CZ.xlsx",
                f"Odchylky_{dd:02d}_{mm:02d}_{yyyy}_V1_CZ.xlsx",
                f"Odchylky_{dd:02d}_{mm:02d}_{yyyy}_V2_CZ.xlsx",
                f"Odchylky_{dd:02d}_{mm:02d}_{yyyy}_V0_CZ.xls",
            ]
            base_path = f"https://www.ote-cr.cz/pubweb/attachments/05_09_12/{yyyy}/month{mm:02d}/day{dd:02d}"
            
            urls = []
            for fn in file_names:
                urls.append(f"{base_path}/{fn}")
                urls.append(f"{base_path}/{fn}/view")
            
            xlsx_bytes = None
            tried_urls = []
            for url in urls:
                tried_urls.append(url)
                try:
                    r = requests.get(url, timeout=15, headers={
                        "User-Agent": "Mozilla/5.0",
                        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, */*",
                    })
                    if r.status_code == 200 and r.content[:2] == b"PK":
                        xlsx_bytes = r.content
                        print(f"  -> OTE ZO {date_str_iso}: XLSX nalezeno @ {url}", flush=True)
                        break
                except Exception as e:
                    print(f"  -> OTE ZO {date_str_iso}: chyba {url}: {e}", flush=True)
                    continue
            
            if xlsx_bytes is None:
                out = {
                    "date": date_str_iso,
                    "qh": [],
                    "count": 0,
                    "error": "no XLSX found",
                    "tried": tried_urls[:3],
                    "fetched_at": datetime.now().isoformat(),
                }
                cache["ts"] = now_ts; cache["data"] = out
                self._json(out); return
            
            # Parse XLSX
            from openpyxl import load_workbook
            from io import BytesIO
            import re as _re
            
            wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
            
            qh_data = []
            KURZ = 24.5
            
            # Pro debug - sber sample obsah z prvních řádků každého sheetu
            debug_sheets = []
            
            # Procházej VŠECHNY sheety, ne jen active
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows: continue
                
                # Debug: ulož prvních 8 řádků 
                sheet_dbg = {
                    "sheet": ws_name,
                    "rows": len(all_rows),
                    "first_rows": []
                }
                for r in all_rows[:8]:
                    if r:
                        sheet_dbg["first_rows"].append([str(c)[:50] if c is not None else None for c in r[:10]])
                debug_sheets.append(sheet_dbg)
                
                # Najdi sloupce - interval (HH:MM-HH:MM nebo Perioda) + cena
                # Strategie: scan VŠECHNY header buňky, ulož kandidaty,
                # pak vyber NEJLEPŠÍ interval + cena ROZDÍLNÉ sloupce
                
                header_row_idx = None
                interval_candidates = []  # [(ri, ci, score), ...]
                price_candidates = []
                currency = "CZK"
                
                for ri, row in enumerate(all_rows[:25]):
                    if not row: continue
                    for ci, val in enumerate(row):
                        if val is None: continue
                        sval = str(val).lower().strip()
                        if not sval: continue
                        
                        # Interval kandidáti - upřednost je explicitní "interval" nebo "od/do"
                        interval_score = 0
                        if "časový interval" in sval or "casovy interval" in sval:
                            interval_score = 100
                        elif sval == "interval" or "interval " in sval or sval.startswith("interval"):
                            interval_score = 10
                        elif "od/do" in sval or "od-do" in sval:
                            interval_score = 9
                        elif sval == "čas" or sval == "cas" or sval == "time":
                            interval_score = 7
                        elif sval == "perioda":
                            interval_score = 2
                        if interval_score:
                            interval_candidates.append((ri, ci, interval_score))
                            if header_row_idx is None: header_row_idx = ri
                        
                        # Cena - hledej KONKRÉTNĚ "zúčtovací cena odchylky"
                        # NE protiodchylky, NE komponenty, NE neuskutečněné aktivace
                        price_score = 0
                        # Nejvyšší priorita: přesný název
                        if "zúčtovací cena odchylky" in sval or "zuctovaci cena odchylky" in sval:
                            price_score = 100  # nejvyšší
                            if "eur" in sval: currency = "EUR"
                            elif "kč" in sval or "czk" in sval: currency = "CZK"
                        # Druhá priorita: jiné varianty zúčtovací ceny
                        elif "zúčtovac" in sval and "cena" in sval and "protiodchyl" not in sval and "komponent" not in sval:
                            price_score = 50
                            if "eur" in sval: currency = "EUR"
                            elif "kč" in sval or "czk" in sval: currency = "CZK"
                        # Pak ostatní (pokud nic lepšího není)
                        elif "cena" in sval and "odchyl" in sval and "protiodchyl" not in sval and "komponent" not in sval:
                            price_score = 10
                            if "eur" in sval: currency = "EUR"
                            elif "kč" in sval or "czk" in sval: currency = "CZK"
                        if price_score:
                            price_candidates.append((ri, ci, price_score))
                            if header_row_idx is None: header_row_idx = ri
                
                # Vyber nejlepší kandidáty (NESMÍ být stejný sloupec)
                interval_col = None
                price_col = None
                
                # Seřaď podle score
                interval_candidates.sort(key=lambda x: (-x[2], x[1]))
                price_candidates.sort(key=lambda x: (-x[2], x[1]))
                
                # Vezmi nejlepší cenu
                if price_candidates:
                    price_col = price_candidates[0][1]
                    if header_row_idx is None: header_row_idx = price_candidates[0][0]
                
                # Vezmi nejlepší interval co NENÍ price_col
                for ri_c, ci_c, sc in interval_candidates:
                    if ci_c != price_col:
                        interval_col = ci_c
                        if header_row_idx is None: header_row_idx = ri_c
                        break
                
                if header_row_idx is None or interval_col is None or price_col is None:
                    continue
                
                # KEY CHECK: interval_col MUSÍ být JINÝ než price_col
                if interval_col == price_col:
                    continue
                
                # Parse data
                for ri in range(header_row_idx + 1, len(all_rows)):
                    row = all_rows[ri]
                    if not row or len(row) <= max(interval_col, price_col): continue
                    
                    iv = row[interval_col]
                    pv = row[price_col]
                    if iv is None or pv is None: continue
                    
                    # Try parse interval as "HH:MM" or "HH:MM-HH:MM" or just hour number
                    interval_str = str(iv) if not isinstance(iv, str) else iv
                    
                    h = None; mi = None
                    m = _re.search(r"(\d{1,2}):(\d{2})", interval_str)
                    if m:
                        h = int(m.group(1)); mi = int(m.group(2))
                    else:
                        # Možná jen hodina jako číslo (1, 2, ..., 96 pro QH index)
                        try:
                            num = int(float(interval_str))
                            if 1 <= num <= 96:
                                # QH index (1-96) → hour, minute
                                idx0 = num - 1
                                h = idx0 // 4
                                mi = (idx0 % 4) * 15
                            elif 0 <= num <= 23:
                                h = num
                                mi = 0
                        except (ValueError, TypeError):
                            continue
                    
                    if h is None or h > 23 or mi is None: continue
                    
                    # Cena
                    try:
                        price_num = float(pv)
                    except (ValueError, TypeError):
                        continue
                    
                    # Currency conversion
                    if currency == "CZK" or abs(price_num) > 500:
                        price_eur = price_num / KURZ
                    else:
                        price_eur = price_num
                    
                    qh_data.append({
                        "hour": h,
                        "minute": mi,
                        "interval": f"{h:02d}:{mi:02d}-{(h + (1 if mi==45 else 0)):02d}:{(mi+15)%60:02d}",
                        "price_eur": round(price_eur, 2),
                        "price_kc": round(price_eur * KURZ, 2),
                    })
                
                if qh_data:
                    # Našli jsme data v tomto sheetu, neprocházet další
                    break
            
            # Dedupe podle (hour, minute)
            seen = set()
            deduped = []
            for r in qh_data:
                k = (r["hour"], r["minute"])
                if k not in seen:
                    seen.add(k)
                    deduped.append(r)
            deduped.sort(key=lambda r: (r["hour"], r["minute"]))
            qh_data = deduped
            
            out = {
                "date": date_str_iso,
                "day_offset": int(qs.get("day", ["?"])[0]) if not date_param else None,
                "qh": qh_data,
                "count": len(qh_data),
                "source": "ote-cr.cz Odchylky XLSX",
                "fetched_at": datetime.now().isoformat(),
            }
            
            # Sanity check - pokud price_kc tvoří sekvenci (1,2,3,4...), 
            # parser sebral sloupec PERIODA místo CENY
            if len(qh_data) >= 10:
                sample_prices = [r["price_kc"] for r in qh_data[:10]]
                # Test sekvenčnosti
                diffs = [sample_prices[i+1] - sample_prices[i] for i in range(len(sample_prices)-1)]
                if all(0.9 < d < 1.1 for d in diffs):
                    # Vyresetuj - parser vzal špatný sloupec
                    qh_data = []
            
            # Pokud parsing selhal nebo malé data, přilož debug info
            if len(qh_data) < 80:  # méně než 80 QH = něco špatně
                out["debug_sheets"] = debug_sheets[:3]
                if not qh_data:
                    out["error"] = "parser found no QH data"
                else:
                    out["warning"] = f"parsed only {len(qh_data)} QH (expected 96)"
            
            cache["ts"] = now_ts; cache["data"] = out
            self._json(out); return
            
        except Exception as e:
            import traceback
            self._json({
                "error": f"OTE ZO: {e}",
                "trace": traceback.format_exc()[:500]
            }, 500)

    def _ote_qh(self, qs):
        """Vraci 15-min data ze spotovaelektrina.cz get-prices-json-qh.
        Vystup:
          last_8: 8 poslednich ctvrthodin (2h dozadu vc. probihajici)
          next_8: 8 nasledujicich ctvrthodin (2h dopredu)
        Cache 5 min.
        """
        try:
            nocache = qs.get("nocache", ["0"])[0] in ("1", "true", "yes")
            if "_OTE_QH_CACHE" not in globals():
                globals()["_OTE_QH_CACHE"] = {"ts": 0, "data": None}
            cache = globals()["_OTE_QH_CACHE"]
            now = time.time()
            if not nocache and cache["data"] and (now - cache["ts"]) < 300:
                out = dict(cache["data"]); out["_cache"] = "hit"
                self._json(out); return

            r = _request_with_retry(
                requests.get,
                "https://spotovaelektrina.cz/api/v1/price/get-prices-json-qh",
                timeout=15,
                headers={"User-Agent": "Mozilla/5.0 (compatible; ceps-dashboard)"}
            )
            if r.status_code != 200:
                self._json({"error": f"OTE QH HTTP {r.status_code}"}, 502); return

            data = r.json()
            hours_today = data.get("hoursToday", [])
            hours_tomorrow = data.get("hoursTomorrow", [])

            # Spocitej current QH index v Berlin TZ
            now_utc = datetime.now(timezone.utc)
            month_now = now_utc.month
            berlin_offset = 2 if 4 <= month_now <= 10 else 1
            berlin_now = now_utc + timedelta(hours=berlin_offset)
            today_date = berlin_now.strftime("%Y-%m-%d")
            tomorrow_date = (berlin_now + timedelta(days=1)).strftime("%Y-%m-%d")
            yesterday_date = (berlin_now - timedelta(days=1)).strftime("%Y-%m-%d")

            current_hour = berlin_now.hour
            current_minute_qh = (berlin_now.minute // 15) * 15  # 0, 15, 30, 45

            # Pomocna funkce na dohledani QH v poli
            def find_qh(arr, h, m):
                return next((row for row in arr
                            if row.get("hour") == h and row.get("minute") == m), None)

            # Funkce pro vraceni cele 15-min strukturky (vc. datumu a noveho casu)
            def make_qh(date_str, h, m, row, is_current=False):
                if row and row.get("priceEur") is not None:
                    return {
                        "date": date_str,
                        "hour": h,
                        "minute": m,
                        "price_eur": round(float(row.get("priceEur")), 2),
                        "price_czk": row.get("priceCZK"),
                        "is_current": is_current,
                    }
                return {
                    "date": date_str,
                    "hour": h,
                    "minute": m,
                    "price_eur": None,
                    "price_czk": None,
                    "is_current": is_current,
                }

            # === LAST 8 (vc. probihajici QH) ===
            # Posledni 8 ctvrthodin koncici aktualni
            last_8 = []
            for offset in range(-7, 1):
                # Vypocitej target QH (h, m, date)
                total_qh_now = current_hour * 4 + (current_minute_qh // 15)
                target_qh_idx = total_qh_now + offset  # muze byt zaporne
                # Konvertuj na (date, hour, minute)
                if target_qh_idx >= 0 and target_qh_idx < 96:
                    # Dnesek
                    th = target_qh_idx // 4
                    tm = (target_qh_idx % 4) * 15
                    row = find_qh(hours_today, th, tm)
                    last_8.append(make_qh(today_date, th, tm, row, is_current=(offset == 0)))
                elif target_qh_idx < 0:
                    # Vcerejsek - zatim None, dosadime z ENTSO-E nebo history
                    yh = (96 + target_qh_idx) // 4
                    ym = ((96 + target_qh_idx) % 4) * 15
                    last_8.append(make_qh(yesterday_date, yh, ym, None, is_current=False))

            # === NEXT 8 (od ted+15 dale) ===
            next_8 = []
            for offset in range(1, 9):
                total_qh_now = current_hour * 4 + (current_minute_qh // 15)
                target_qh_idx = total_qh_now + offset
                if target_qh_idx < 96:
                    th = target_qh_idx // 4
                    tm = (target_qh_idx % 4) * 15
                    row = find_qh(hours_today, th, tm)
                    next_8.append(make_qh(today_date, th, tm, row))
                else:
                    # Zitrek
                    excess = target_qh_idx - 96
                    th = excess // 4
                    tm = (excess % 4) * 15
                    row = find_qh(hours_tomorrow, th, tm)
                    next_8.append(make_qh(tomorrow_date, th, tm, row))

            # Cele pole pro tabulku - dnesni + zitrejsi
            qh_today = []
            for row in hours_today:
                if row.get("priceEur") is not None:
                    qh_today.append({
                        "hour": row.get("hour"),
                        "minute": row.get("minute", 0),
                        "price_eur": round(float(row.get("priceEur")), 2),
                        "price_czk": row.get("priceCZK"),
                    })
            qh_tomorrow = []
            for row in hours_tomorrow:
                if row.get("priceEur") is not None:
                    qh_tomorrow.append({
                        "hour": row.get("hour"),
                        "minute": row.get("minute", 0),
                        "price_eur": round(float(row.get("priceEur")), 2),
                        "price_czk": row.get("priceCZK"),
                    })

            out = {
                "last_8": last_8,
                "next_8": next_8,
                "qh_today": qh_today,
                "qh_tomorrow": qh_tomorrow,
                "tomorrow_published": len(hours_tomorrow) > 0,
                "current_hour": current_hour,
                "current_minute_qh": current_minute_qh,
                "source": "spotovaelektrina.cz (15-min OTE)",
                "fetched_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                "_cache": "miss",
            }
            cache["ts"] = now
            cache["data"] = {k: v for k, v in out.items() if k != "_cache"}
            self._json(out)
        except Exception as e:
            print(f"  -> /ote/qh ERROR: {e}", flush=True)
            import traceback
            traceback.print_exc()
            self._json({"error": str(e)}, 502)

    def _ote_spot(self, qs):
        """Vraci aktualni spotovou cenu + statistiky pro cely den.
        Cache 5 minut. Vystup: {price_czk, price_eur, hour, day_stats: {...}}
        """
        try:
            nocache = qs.get("nocache", ["0"])[0] in ("1", "true", "yes")
            # Cache - drzi se 5 minut
            if "_OTE_SPOT_CACHE" not in globals():
                globals()["_OTE_SPOT_CACHE"] = {"ts": 0, "data": None}
            # Historie hodinovych cen napric dny - format: {"2026-05-06": [{"hour": 0, "priceEur": ...}, ...], ...}
            if "_OTE_HISTORY" not in globals():
                globals()["_OTE_HISTORY"] = {}
            cache = globals()["_OTE_SPOT_CACHE"]
            history = globals()["_OTE_HISTORY"]
            now = time.time()
            if not nocache and cache["data"] and (now - cache["ts"]) < 300:
                out = dict(cache["data"]); out["_cache"] = "hit"
                self._json(out); return

            # 1) Stahni aktualni cenu (pro hour info)
            r1 = _request_with_retry(
                requests.get,
                "https://spotovaelektrina.cz/api/v1/price/get-actual-price-json",
                timeout=15,
                headers={"User-Agent": "Mozilla/5.0 (compatible; ceps-dashboard)"}
            )
            if r1.status_code != 200:
                self._json({"error": f"OTE actual HTTP {r1.status_code}"}, 502); return
            actual = r1.json()

            current_hour = actual.get("hour")
            current_eur = actual.get("priceEUR")
            current_czk = actual.get("priceCZK")

            # API neposkytuje "hour" klic - vezmeme aktualni hodinu z Berlin timezone
            if current_hour is None:
                now_utc = datetime.now(timezone.utc)
                # Berlin = CET/CEST (zima UTC+1, leto UTC+2)
                month = now_utc.month
                berlin_offset = 2 if 4 <= month <= 10 else 1
                berlin_now = now_utc + timedelta(hours=berlin_offset)
                current_hour = berlin_now.hour

            # Initialize
            day_stats = None
            tomorrow_stats = None
            last_8h = None  # 8 poslednich hodin pro mini KPI bunky
            hours_tomorrow = []  # zitrejsi ceny (po 14:00 CET)

            # 2) Stahni 24h ceny pro statistiky
            try:
                r2 = _request_with_retry(
                    requests.get,
                    "https://spotovaelektrina.cz/api/v1/price/get-prices-json",
                    timeout=15,
                    headers={"User-Agent": "Mozilla/5.0 (compatible; ceps-dashboard)"}
                )
                if r2.status_code == 200:
                    day_data = r2.json()
                    # Format: {"hoursToday": [{"hour": 0, "priceCZK":..., "priceEur":...}, ...], "hoursTomorrow": [...]}
                    # POZOR: API vraci "priceEur" (male r), ne "priceEUR"!
                    hours_today = day_data.get("hoursToday", [])
                    hours_yesterday = []  # spotovaelektrina.cz nevraci

                    # Spocitej dnesni datum (Berlin TZ)
                    now_utc = datetime.now(timezone.utc)
                    month_now = now_utc.month
                    berlin_offset = 2 if 4 <= month_now <= 10 else 1
                    berlin_now = now_utc + timedelta(hours=berlin_offset)
                    today_date_str = berlin_now.strftime("%Y-%m-%d")
                    yesterday_date_str = (berlin_now - timedelta(days=1)).strftime("%Y-%m-%d")

                    # ULOZ dnesni data do historie - uchovavame max 7 dni zpetne
                    if hours_today:
                        history[today_date_str] = hours_today
                        # Vyciste stara data starsi nez 7 dni
                        cutoff = (berlin_now - timedelta(days=7)).strftime("%Y-%m-%d")
                        for old_date in list(history.keys()):
                            if old_date < cutoff:
                                del history[old_date]

                    # Vcerejsek z historie (kdyz uz ho mame z drivejska)
                    if yesterday_date_str in history:
                        hours_yesterday = history[yesterday_date_str]
                        print(f"  -> OTE history: vcerejsek ({yesterday_date_str}) z pameti, {len(hours_yesterday)}h", flush=True)

                    # Fallback pro vcerejsek - stahnout z ENTSO-E (uz mame token, stejna API jako solar)
                    if not hours_yesterday and current_hour is not None and current_hour < 8:
                        try:
                            yesterday_dt = berlin_now - timedelta(days=1)
                            ps = yesterday_dt.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(hours=berlin_offset)
                            pe = ps + timedelta(hours=24)
                            entsoe_xml, entsoe_st = call_entsoe({
                                "documentType": "A44",
                                "in_Domain":    CZ_DOMAIN,
                                "out_Domain":   CZ_DOMAIN,
                                "periodStart":  fmt_entsoe_period(ps),
                                "periodEnd":    fmt_entsoe_period(pe),
                            })
                            if entsoe_st == 200:
                                # Parsuj XML - struktura je <Period><Point><position>1</position><price.amount>X</price.amount></Point>...
                                import re as _re
                                body_no_ns = _re.sub(r'\sxmlns="[^"]+"', '', entsoe_xml, count=1)
                                root = ET.fromstring(body_no_ns)
                                # Najdi prvni TimeSeries -> Period -> body
                                hours_yesterday = []
                                for ts_node in root.findall(".//TimeSeries"):
                                    period = ts_node.find("Period")
                                    if period is None: continue
                                    res = period.find("resolution")
                                    if res is None or res.text != "PT60M": continue
                                    for pt in period.findall("Point"):
                                        pos_el = pt.find("position")
                                        price_el = pt.find("price.amount")
                                        if pos_el is None or price_el is None: continue
                                        try:
                                            pos = int(pos_el.text)
                                            price = float(price_el.text)
                                            # position 1 = 00:00, position 24 = 23:00
                                            hr = pos - 1
                                            if 0 <= hr <= 23:
                                                hours_yesterday.append({
                                                    "hour": hr,
                                                    "priceEur": price,
                                                    "priceCZK": int(price * 24.35),  # priblizny EUR->CZK
                                                })
                                        except (ValueError, TypeError):
                                            continue
                                    if hours_yesterday:
                                        break  # mame data, ostatni TimeSeries jsou duplicates
                                if hours_yesterday:
                                    print(f"  -> ENTSO-E A44: {len(hours_yesterday)}h vcerejsich CZ cen", flush=True)
                                    history[yesterday_date_str] = hours_yesterday
                            else:
                                print(f"  -> ENTSO-E A44 status {entsoe_st}", flush=True)
                        except Exception as e:
                            print(f"  -> ENTSO-E A44 fallback failed: {e}", flush=True)

                    if hours_today:
                        prices_eur = [h.get("priceEur") for h in hours_today if h.get("priceEur") is not None]
                        if prices_eur:
                            min_eur = min(prices_eur)
                            max_eur = max(prices_eur)
                            avg_eur = sum(prices_eur) / len(prices_eur)
                            min_hour = next((h["hour"] for h in hours_today if h.get("priceEur") == min_eur), None)
                            max_hour = next((h["hour"] for h in hours_today if h.get("priceEur") == max_eur), None)
                            spread = max_eur - min_eur
                            # Vs prumer pro aktualni hodinu
                            vs_avg_pct = None
                            if current_eur is not None and avg_eur > 0:
                                vs_avg_pct = ((current_eur - avg_eur) / avg_eur) * 100
                            day_stats = {
                                "min_eur": round(min_eur, 2),
                                "max_eur": round(max_eur, 2),
                                "avg_eur": round(avg_eur, 2),
                                "min_hour": min_hour,
                                "max_hour": max_hour,
                                "spread_eur": round(spread, 2),
                                "current_vs_avg_pct": round(vs_avg_pct, 1) if vs_avg_pct is not None else None,
                            }
                    # Zitrejsi statistiky (pokud OTE uz publikovalo - obvykle po 14:00 CET)
                    hours_tomorrow = day_data.get("hoursTomorrow", [])
                    
                    # PRIMARNI ZDROJ pro zitra: OTE-CR HTML scrape (60min ceny z posledniho sloupce)
                    try:
                        now_utc = datetime.now(timezone.utc)
                        month = now_utc.month
                        berlin_offset = 2 if 4 <= month <= 10 else 1
                        berlin_now = now_utc + timedelta(hours=berlin_offset)
                        tomorrow_date = (berlin_now + timedelta(days=1)).strftime("%Y-%m-%d")
                        
                        # HTML stranka OTE - obsahuje tabulku se sloupci:
                        # | Casovy interval | 15min cena | ... | 60min cena |
                        ote_url = f"https://www.ote-cr.cz/cs/kratkodobe-trhy/elektrina/denni-trh?date={tomorrow_date}&time_resolution=PT15M"
                        ote_r = requests.get(ote_url, timeout=15, headers={
                            "User-Agent": "Mozilla/5.0 (compatible; ceps-dashboard)",
                            "Accept": "text/html",
                        })
                        if ote_r.status_code == 200:
                            html = ote_r.text
                            ote_tomorrow = []
                            import re as _re
                            matched_lines = 0
                            for line in html.split("\n"):
                                # Hledame radky s "HH:00-HH:15" kdekoliv v radku
                                m = _re.search(r"(\d{2}):00-(\d{2}):15", line)
                                if not m:
                                    continue
                                matched_lines += 1
                                hour = int(m.group(1))
                                # Najdi vsechny ceny X,XX v radku (vcetne zapornych a s mezerou jako oddelovac tisicu)
                                nums = _re.findall(r"-?\d+(?:[\s\u00A0]\d+)*,\d+", line)
                                if not nums:
                                    continue
                                # Posledni cislo = 60min cena (posledni sloupec)
                                try:
                                    price_str = nums[-1].replace(" ", "").replace("\u00A0", "").replace(",", ".")
                                    price = float(price_str)
                                    if 0 <= hour <= 23 and price != 0:
                                        ote_tomorrow.append({"hour": hour, "priceEur": price})
                                except: pass
                            
                            print(f"  -> OTE-CR tomorrow scrape: HTML {len(html)}b, matched lines={matched_lines}, parsed={len(ote_tomorrow)} h", flush=True)
                            if len(ote_tomorrow) >= 20:  # ocekavame 24 hodin
                                hours_tomorrow = ote_tomorrow
                    except Exception as e:
                        print(f"  -> OTE-CR tomorrow fallback: {e}", flush=True)
                    
                    if hours_tomorrow:
                        prices_t = [h.get("priceEur") for h in hours_tomorrow if h.get("priceEur") is not None]
                        if prices_t:
                            tmin = min(prices_t)
                            tmax = max(prices_t)
                            tavg = sum(prices_t) / len(prices_t)
                            tmin_h = next((h["hour"] for h in hours_tomorrow if h.get("priceEur") == tmin), None)
                            tmax_h = next((h["hour"] for h in hours_tomorrow if h.get("priceEur") == tmax), None)
                            tomorrow_stats = {
                                "min_eur": round(tmin, 2),
                                "max_eur": round(tmax, 2),
                                "avg_eur": round(tavg, 2),
                                "min_hour": tmin_h,
                                "max_hour": tmax_h,
                                "spread_eur": round(tmax - tmin, 2),
                                "published": True,
                            }
                        else:
                            tomorrow_stats = {"published": False}
                    else:
                        tomorrow_stats = {"published": False}

                    # last_8h: VZDY posledni 8 hodin koncici aktualni hodinou
                    # ch=3 -> [yesterday-20, yesterday-21, yesterday-22, yesterday-23, today-0, today-1, today-2, today-3]
                    if current_hour is not None and hours_today:
                        ch = current_hour
                        last_8h_list = []
                        for offset in range(-7, 1):
                            target_hour = ch + offset
                            is_yesterday = target_hour < 0  # FIX: dle target_hour, NE offsetu
                            if not is_yesterday:
                                # Aktualni den
                                row = next((h for h in hours_today if h.get("hour") == target_hour), None)
                                display_hour = target_hour
                            else:
                                # Predchozi den
                                yesterday_hour = target_hour + 24
                                row = next((h for h in hours_yesterday if h.get("hour") == yesterday_hour), None)
                                display_hour = yesterday_hour
                            if row and row.get("priceEur") is not None:
                                last_8h_list.append({
                                    "hour": display_hour,
                                    "price_eur": round(row.get("priceEur"), 2),
                                    "price_czk": row.get("priceCZK"),
                                    "is_current": offset == 0,
                                    "is_yesterday": is_yesterday,
                                })
                            else:
                                last_8h_list.append({
                                    "hour": display_hour,
                                    "price_eur": None,
                                    "price_czk": None,
                                    "is_current": offset == 0,
                                    "is_yesterday": is_yesterday,
                                })
                        last_8h = last_8h_list
            except Exception as e:
                print(f"  -> OTE day stats fetch failed: {e}", flush=True)

            out = {
                "price_czk": current_czk,
                "price_eur": current_eur,
                "hour": current_hour,
                "day_stats": day_stats,
                "tomorrow_stats": tomorrow_stats,
                "last_8h": last_8h,
                "hours_today": hours_today,  # vsechny dnesni hodinove ceny
                "hours_tomorrow": hours_tomorrow,  # zitrejsi po 14:00 CET
                "source": "spotovaelektrina.cz",
                "fetched_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                "_cache": "miss",
            }
            cache["ts"] = now
            cache["data"] = {k: v for k, v in out.items() if k != "_cache"}
            self._json(out)
        except Exception as e:
            print(f"  -> OTE Spot ERROR: {e}", flush=True)
            self._json({"error": str(e)}, 502)

    def _fetch_wttr(self, location):
        """Stahne pocasi z wttr.in pro jednu lokaci. Vraci dict nebo None.
        Format vystupu sjednocen s _weather strukturou.
        """
        # wttr.in weather codes -> emoji + cesky popis
        def wcode_info(code):
            if not code: return ("", "—")
            c = str(code)
            if c in ("113",):                       return ("☀️", "jasno")
            if c in ("116",):                       return ("🌤️", "polojasno")
            if c in ("119", "122"):                 return ("☁️", "zataženo")
            if c in ("143", "248", "260"):          return ("🌫️", "mlha")
            if c in ("176", "263", "266", "281", "284", "293", "296"): return ("🌦️", "mrholení")
            if c in ("299", "302", "305", "308", "311", "314", "353", "356", "359"): return ("🌧️", "déšť")
            if c in ("179", "182", "185", "227", "230", "317", "320",
                     "323", "326", "329", "332", "335", "338", "350",
                     "362", "365", "368", "371", "374", "377"):        return ("🌨️", "sněžení")
            if c in ("386", "389"):                 return ("⛈️", "bouřka")
            if c in ("392", "395"):                 return ("❄️", "sněhové bouřky")
            if c in ("200",):                       return ("⛈️", "bouřka")
            return ("", "—")

        try:
            r = _request_with_retry(
                requests.get, f"https://wttr.in/{location}",
                params={"format": "j1"}, timeout=15,
                headers={"User-Agent": "Mozilla/5.0 (compatible; ceps-dashboard)"}
            )
            if r.status_code != 200:
                return None
            data = r.json()
        except Exception as e:
            print(f"  -> wttr.in fetch fail ({location}): {e}", flush=True)
            return None

        try:
            cur = data["current_condition"][0]
            today = data["weather"][0]
            tom = data["weather"][1] if len(data["weather"]) > 1 else None

            # Current
            cur_temp = float(cur.get("temp_C", 0))
            cur_wind_kmh = float(cur.get("windspeedKmph", 0))
            cur_wind_ms = round(cur_wind_kmh / 3.6, 1)
            cur_code = cur.get("weatherCode")
            cur_cloud = int(cur.get("cloudcover", 0))

            # Today
            t_max = float(today.get("maxtempC", 0))
            t_min = float(today.get("mintempC", 0))
            # Wind max - hourly data, najdi nejvyssi
            t_winds_kmh = [float(h.get("windspeedKmph", 0)) for h in today.get("hourly", [])]
            t_wind_max_ms = round(max(t_winds_kmh) / 3.6, 1) if t_winds_kmh else None
            # Sunshine - wttr.in nedava sunshine_duration, pouzijem sunHour
            t_sun = float(today.get("sunHour", 0))
            t_codes = [h.get("weatherCode") for h in today.get("hourly", [])]
            # Vezmi midday code (12:00 = idx 4 u 3h intervals)
            t_code = t_codes[4] if len(t_codes) > 4 else (t_codes[0] if t_codes else None)
            t_icon, t_desc = wcode_info(t_code)

            # Tomorrow
            if tom:
                tom_max = float(tom.get("maxtempC", 0))
                tom_min = float(tom.get("mintempC", 0))
                tom_winds_kmh = [float(h.get("windspeedKmph", 0)) for h in tom.get("hourly", [])]
                tom_wind_max_ms = round(max(tom_winds_kmh) / 3.6, 1) if tom_winds_kmh else None
                tom_sun = float(tom.get("sunHour", 0))
                tom_codes = [h.get("weatherCode") for h in tom.get("hourly", [])]
                tom_code = tom_codes[4] if len(tom_codes) > 4 else (tom_codes[0] if tom_codes else None)
                tom_icon, tom_desc = wcode_info(tom_code)
            else:
                tom_max = tom_min = tom_wind_max_ms = tom_sun = None
                tom_icon, tom_desc = ("", "—")

            return {
                "current": {
                    "temp_c": cur_temp,
                    "wind_ms": cur_wind_ms,
                    "cloud_pct": cur_cloud,
                    "weather_code": cur_code,
                },
                "today": {
                    "temp_max": t_max,
                    "temp_min": t_min,
                    "wind_max_ms": t_wind_max_ms,
                    "icon": t_icon,
                    "desc": t_desc,
                    "sunshine_h": t_sun,
                },
                "tomorrow": {
                    "temp_max": tom_max,
                    "temp_min": tom_min,
                    "wind_max_ms": tom_wind_max_ms,
                    "icon": tom_icon,
                    "desc": tom_desc,
                    "sunshine_h": tom_sun,
                },
            }
        except Exception as e:
            print(f"  -> wttr.in parse fail ({location}): {e}", flush=True)
            return None

    def _weather(self, qs):
        """Vraci pocasi pro Prahu z wttr.in.
        Cache 3 hodiny + stale-while-revalidate.
        """
        try:
            if "_WEATHER_CACHE" not in globals():
                globals()["_WEATHER_CACHE"] = {"ts": 0, "data": None}
            cache = globals()["_WEATHER_CACHE"]
            now = time.time()
            if cache["data"] and (now - cache["ts"]) < 10800:
                out = dict(cache["data"]); out["_cache"] = "hit"
                out["_age_sec"] = int(now - cache["ts"])
                self._json(out); return

            result = self._fetch_wttr("Praha")
            if result is None:
                # Stale fallback
                if cache["data"]:
                    out = dict(cache["data"])
                    out["_cache"] = "stale"
                    out["_age_sec"] = int(now - cache["ts"])
                    self._json(out); return
                self._json({"error": "wttr.in failed"}, 200); return

            out = result
            out["location"] = "Praha"
            out["source"] = "wttr.in"
            out["fetched_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
            out["_cache"] = "miss"
            cache["ts"] = now
            cache["data"] = {k: v for k, v in out.items() if k != "_cache"}
            self._json(out)
        except Exception as e:
            print(f"  -> Weather ERROR: {e}", flush=True)
            self._json({"error": str(e)}, 502)




    # ----------------------------------------------------------------
    # Netztransparenz.de - Aktivierte aFRR + mFRR + NRV-Saldo DE
    # Endpointy: /netztransparenz/activated?type=afrr|mfrr|nrv&date=YYYY-MM-DD
    # Cache 15min, OAuth2 client_credentials token cache 55min
    # ----------------------------------------------------------------
    _NTP_TOKEN_CACHE = {"token": None, "expires": 0}

    def _ntp_get_token(self):
        import urllib.request, urllib.parse, urllib.error
        now = time.time()
        if self._NTP_TOKEN_CACHE["token"] and now < self._NTP_TOKEN_CACHE["expires"]:
            return self._NTP_TOKEN_CACHE["token"]
        client_id = os.environ.get("NTP_CLIENT_ID", "").strip()
        client_secret = os.environ.get("NTP_CLIENT_SECRET", "").strip()
        if not client_id or not client_secret:
            raise Exception("NTP_CLIENT_ID / NTP_CLIENT_SECRET not configured")
        data = urllib.parse.urlencode({
            "grant_type": "client_credentials",
            "client_id": client_id,
            "client_secret": client_secret,
        }).encode("utf-8")
        req = urllib.request.Request(
            "https://identity.netztransparenz.de/users/connect/token",
            data=data,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=15) as r:
            tok = json.loads(r.read().decode("utf-8"))
        token = tok["access_token"]
        expires_in = int(tok.get("expires_in", 3600))
        Handler._NTP_TOKEN_CACHE["token"] = token
        Handler._NTP_TOKEN_CACHE["expires"] = now + expires_in - 60
        print(f"  -> NTP token OK, expires in {expires_in}s", flush=True)
        return token

    def _ntp_activated(self, qs):
        import urllib.request, urllib.error, io, csv
        try:
            typ = qs.get("type", ["afrr"])[0].lower()
            nocache = qs.get("nocache", ["0"])[0] in ("1","true","yes")

            # Datum: CET dnes
            now_utc = datetime.now(timezone.utc)
            berlin_off = 2 if 4 <= now_utc.month <= 10 else 1
            berlin_now = now_utc + timedelta(hours=berlin_off)
            date_str = qs.get("date", [berlin_now.strftime("%Y-%m-%d")])[0]

            # Endpunkt podle typu
            endpoint_map = {
                "afrr": "NrvSaldo/AktivierteSRL/Betrieblich",
                "mfrr": "NrvSaldo/AktivierteMRL/Betrieblich",
                "nrv":  "NrvSaldo/NRVSaldo/Betrieblich",
            }
            if typ not in endpoint_map:
                self._json({"error": f"Unknown type: {typ}, use afrr|mfrr|nrv"}, 400); return

            cache_key = f"_NTP_{typ.upper()}_{date_str}"
            if cache_key not in globals():
                globals()[cache_key] = {"ts": 0, "data": None}
            cache = globals()[cache_key]
            now_ts = time.time()
            if not nocache and cache["data"] and (now_ts - cache["ts"]) < 900:
                out = dict(cache["data"]); out["_cache"] = "hit"
                self._json(out); return

            # Datum range: CET den -> UTC
            date_from = f"{date_str}T00:00:00"
            date_to   = f"{date_str}T23:59:59"
            # berlin -> utc
            date_from_utc = (datetime.strptime(date_from, "%Y-%m-%dT%H:%M:%S")
                             - timedelta(hours=berlin_off)).strftime("%Y-%m-%dT%H:%M:%S")
            date_to_utc   = (datetime.strptime(date_to,   "%Y-%m-%dT%H:%M:%S")
                             - timedelta(hours=berlin_off)).strftime("%Y-%m-%dT%H:%M:%S")

            token = self._ntp_get_token()
            endpoint = endpoint_map[typ]
            url = (f"https://ds.netztransparenz.de/api/v1/data/{endpoint}"
                   f"?dateFrom={date_from_utc}&dateTo={date_to_utc}")
            print(f"  -> NTP {typ} fetch: {url}", flush=True)
            req = urllib.request.Request(url, headers={
                "Authorization": f"Bearer {token}",
                "Accept": "text/csv",
            })
            with urllib.request.urlopen(req, timeout=20) as r:
                raw_csv = r.read().decode("utf-8-sig")

            # Parsuj CSV Format 6.a nebo 7.a
            rows = []
            reader = csv.DictReader(io.StringIO(raw_csv), delimiter=";")
            for row in reader:
                # Datum;Zeitzone;von;bis;...;Deutschland (Positiv);...;Deutschland (Negativ)
                von = row.get("von","").strip()
                bis = row.get("bis","").strip()
                if not von: continue
                # Konvertuj UTC -> CET
                try:
                    h_utc, m_utc = int(von[:2]), int(von[3:5])
                    h_cet = (h_utc + berlin_off) % 24
                    lbl = f"{h_cet:02d}:{m_utc:02d}"
                except:
                    lbl = von

                pos = neg = None
                for k, v in row.items():
                    kl = k.strip()
                    val = v.replace(",",".").strip() if v else ""
                    if not val or val in ("N.A.",""): continue
                    try: fv = float(val)
                    except: continue
                    if "Deutschland" in kl and "Positiv" in kl: pos = round(fv,3)
                    if "Deutschland" in kl and "Negativ" in kl: neg = round(fv,3)
                    if kl == "Deutschland" and typ == "nrv": pos = round(fv,3)

                rows.append({"lbl": lbl, "von_utc": von, "pos": pos, "neg": neg})

            out = {
                "type": typ,
                "date": date_str,
                "rows": rows,
                "n": len(rows),
                "fetched_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                "_cache": "miss"
            }
            cache["ts"] = now_ts
            cache["data"] = {k:v for k,v in out.items() if k != "_cache"}
            print(f"  -> NTP {typ}: {len(rows)} rows", flush=True)
            self._json(out)
        except urllib.error.HTTPError as e:
            body = ""
            try: body = e.read().decode()[:300]
            except: pass
            print(f"  -> NTP ERROR HTTP {e.code}: {body}", flush=True)
            # Token expired? clear cache
            Handler._NTP_TOKEN_CACHE["token"] = None
            self._json({"error": f"HTTP {e.code}", "detail": body, "url": url if 'url' in dir() else ""}, 200)
        except Exception as e:
            import traceback; traceback.print_exc()
            self._json({"error": str(e)}, 502)

    def _eupowerprices_areas(self):
        try:
            import urllib.request, urllib.error
            api_key = os.environ.get("EUPOWERPRICES_API_KEY", "").strip()
            if not api_key:
                self._json({"error": "not configured"}); return
            req = urllib.request.Request(
                "https://api.eupowerprices.com/v1/areas",
                headers={"X-API-Key": api_key})
            with urllib.request.urlopen(req, timeout=10) as r:
                self._json(json.loads(r.read().decode("utf-8")))
        except urllib.error.HTTPError as e:
            body = ""
            try: body = e.read().decode()[:200]
            except: pass
            self._json({"error": f"HTTP {e.code}", "detail": body})
        except Exception as e:
            self._json({"error": str(e)})

    def _eupowerprices_forecast(self, qs):
        """EU Power Prices - hodinovy forecast cen elektriny.
        Cache 1h.
        Query: ?area=CZ (default CZ, mozne DE, AT, SK, PL, HU, ...)
        Vraci: {area, issued_at, points:[{ts, price_eur}], n}
        """
        try:
            import urllib.request, urllib.error
            api_key = os.environ.get("EUPOWERPRICES_API_KEY", "").strip()
            if not api_key:
                self._json({"error": "EUPOWERPRICES_API_KEY not configured"}, 200); return

            area = qs.get("area", ["CZ"])[0].upper()
            nocache = qs.get("nocache", ["0"])[0] in ("1","true","yes")

            cache_key = f"_EUPOWERPRICES_{area}"
            if cache_key not in globals():
                globals()[cache_key] = {"ts": 0, "data": None}
            cache = globals()[cache_key]
            now = time.time()
            if not nocache and cache["data"] and (now - cache["ts"]) < 3600:
                out = dict(cache["data"]); out["_cache"] = "hit"
                out["_age_sec"] = int(now - cache["ts"])
                self._json(out); return

            url = f"https://api.eupowerprices.com/v1/forecasts/{area}/latest"
            req = urllib.request.Request(url, headers={"X-API-Key": api_key})
            try:
                with urllib.request.urlopen(req, timeout=20) as r:
                    raw = json.loads(r.read().decode("utf-8"))
            except urllib.error.HTTPError as e:
                err_body = ""
                try: err_body = e.read().decode("utf-8")[:300]
                except: pass
                self._json({"error": f"HTTP {e.code}", "detail": err_body, "area": area}, 200); return

            # Parsuj odpoved - series field
            points = []
            issued_at = raw.get("generated_at_utc") or raw.get("issued_at") or raw.get("forecast_time") or ""

            series = raw.get("series") or []
            if isinstance(series, list):
                for item in series:
                    if not isinstance(item, dict): continue
                    ts = (item.get("ts_utc") or item.get("datetime_utc") or item.get("datetime")
                          or item.get("ts") or item.get("time"))
                    price = (item.get("price_eur_mwh") or item.get("value") or item.get("price")
                             or item.get("price_eur"))
                    if ts is None or price is None: continue
                    try: price_f = float(price)
                    except: continue
                    points.append({"ts": str(ts), "price_eur": round(price_f, 2)})

            out = {
                "area": area,
                "issued_at": issued_at,
                "points": points,
                "n": len(points),
                "raw_keys": list(raw.keys()) if isinstance(raw, dict) else [],
                "fetched_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                "_cache": "miss"
            }
            cache["ts"] = now
            cache["data"] = {k: v for k, v in out.items() if k != "_cache"}
            print(f"  -> eupowerprices {area}: {len(points)} bodu", flush=True)
            self._json(out)
        except Exception as e:
            print(f"  -> eupowerprices ERROR: {e}", flush=True)
            self._json({"error": str(e)}, 502)

    def _metdesk_magma(self, qs):
        """MetDesk Power Generation V2 - solar/wind forecast pro DE/AT/CZ/HU.
        Cache 1h.
        Query: ?model=magma&country=CZ&generation_type=solar (nebo wind)
        """
        try:
            import urllib.request, urllib.error
            api_key = os.environ.get("METDESK_API_KEY", "")
            # Normalizace: nahrad vsechny newliny/taby/multi-spaces jednou mezerou
            import re
            api_key = re.sub(r'\s+', ' ', api_key).strip()
            if not api_key:
                self._json({"error": "METDESK_API_KEY not configured"}, 200); return

            if "_MAGMA_CACHE" not in globals():
                globals()["_MAGMA_CACHE"] = {}
            
            model = qs.get("model", ["magma"])[0]
            country = qs.get("country", ["CZ"])[0]
            gen_type = qs.get("generation_type", ["solar"])[0]
            
            cache_key = f"{model}|{country}|{gen_type}"
            cache_all = globals()["_MAGMA_CACHE"]
            now = time.time()
            if cache_key in cache_all and (now - cache_all[cache_key]["ts"]) < 3600:
                out = dict(cache_all[cache_key]["data"]); out["_cache"] = "hit"
                out["_age_sec"] = int(now - cache_all[cache_key]["ts"])
                self._json(out); return

            raw_token = api_key.split(" ", 1)[1] if " " in api_key else api_key
            # Pokud uz ma "jwt " prefix v env var, pouzij as-is. Jinak pridej "jwt ".
            lower = api_key.lower()
            if lower.startswith("jwt "):
                auth_hdr = api_key  # uz ma "jwt prefix"
            else:
                auth_hdr = f"jwt {raw_token}"
            
            def fetch(url):
                req = urllib.request.Request(url, headers={"Authorization": auth_hdr})
                with urllib.request.urlopen(req, timeout=20) as r:
                    return json.loads(r.read().decode("utf-8"))
            
            issues_url = f"https://api.metdesk.com/get/metdesk/powergen/v2/issues?model={model}"
            try:
                issues_raw = fetch(issues_url)
            except urllib.error.HTTPError as e:
                err_body = ""
                try: err_body = e.read().decode("utf-8")[:300]
                except: pass
                self._json({"error": f"issues {e.code}", "detail": err_body, "url": issues_url}, 200); return
            
            issues_list = issues_raw.get("data", [])
            if not issues_list:
                self._json({"error": "no issues", "raw": issues_raw}, 200); return
            latest = issues_list[-1]
            latest_issue = latest if isinstance(latest, str) else latest.get("issue", str(latest))
            
            start = datetime.now(timezone.utc).strftime("%Y-%m-%dT00:00:00Z")
            end = (datetime.now(timezone.utc) + timedelta(hours=48)).strftime("%Y-%m-%dT00:00:00Z")
            fc_url = (f"https://api.metdesk.com/get/metdesk/powergen/v2/forecasts"
                      f"?model={model}&issue={latest_issue}&location={country}"
                      f"&location_type=country&element={gen_type}&interval=hires"
                      f"&start_dtg={start}&end_dtg={end}")
            try:
                raw = fetch(fc_url)
            except urllib.error.HTTPError as e:
                err_body = ""
                try: err_body = e.read().decode("utf-8")[:300]
                except: pass
                self._json({"error": f"forecasts {e.code}", "detail": err_body, "url": fc_url}, 200); return
            
            data = raw.get("data", [])
            points = []
            for item in data:
                dtg = item.get("dtg") or item.get("datetime")
                val = item.get("value") or item.get("power")
                if dtg is None or val is None: continue
                try: vf = float(val)
                except: continue
                points.append({"ts": dtg, "value": vf})
            
            out = {
                "model": model,
                "country": country,
                "generation_type": gen_type,
                "issue": latest_issue,
                "points": points,
                "n": len(points),
                "fetched_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                "_cache": "miss"
            }
            cache_all[cache_key] = {"ts": now, "data": {k: v for k, v in out.items() if k != "_cache"}}
            self._json(out)
        except Exception as e:
            print(f"  -> POWER ERROR: {e}", flush=True)
            self._json({"error": str(e)}, 502)

    def _metdesk_weather(self, qs):
        """MetDesk MAGMA Weather API v1 - meteo forecast (vitr, solar rad, teplota...).
        Cache 1h. Forecast se vydava kazdou hodinu, horizont 15 dni.
        Query: ?location=DE&elements=ff_100m,rad,tt&interval=60
          location  - kod lokace (napr. DE, GB, nebo konkretni misto z /locations)
          elements  - carkou oddeleny seznam element kodu (default: ff_100m,rad,tt,nt)
          interval  - 60 | 30 | 15 (default 60)
        Vraci: {location, issue, interval, points:[{ts, <element>:val, ...}], elements:[...]}
        """
        try:
            import urllib.request, urllib.error, re
            api_key = os.environ.get("METDESK_API_KEY", "")
            api_key = re.sub(r'\s+', ' ', api_key).strip()
            if not api_key:
                self._json({"error": "METDESK_API_KEY not configured"}, 200); return

            if "_WEATHER_CACHE" not in globals():
                globals()["_WEATHER_CACHE"] = {}

            location = qs.get("location", ["DE"])[0]
            elements = qs.get("elements", ["ff_100m,rad,tt,nt"])[0]
            interval = qs.get("interval", ["60"])[0]
            el_list = [e.strip() for e in elements.split(",") if e.strip()]

            cache_key = f"{location}|{elements}|{interval}"
            cache_all = globals()["_WEATHER_CACHE"]
            now = time.time()
            if cache_key in cache_all and (now - cache_all[cache_key]["ts"]) < 3600:
                out = dict(cache_all[cache_key]["data"]); out["_cache"] = "hit"
                out["_age_sec"] = int(now - cache_all[cache_key]["ts"])
                self._json(out); return

            raw_token = api_key.split(" ", 1)[1] if " " in api_key else api_key
            lower = api_key.lower()
            auth_hdr = api_key if lower.startswith("jwt ") else f"jwt {raw_token}"

            BASE = "https://api.metdesk.com/get/metdesk/magmaweather/v1"

            def fetch(url):
                req = urllib.request.Request(url, headers={"Authorization": auth_hdr})
                with urllib.request.urlopen(req, timeout=20) as r:
                    return json.loads(r.read().decode("utf-8"))

            # 1) najdi nejnovejsi issue
            issues_url = f"{BASE}/issues"
            try:
                issues_raw = fetch(issues_url)
            except urllib.error.HTTPError as e:
                err_body = ""
                try: err_body = e.read().decode("utf-8")[:300]
                except: pass
                self._json({"error": f"issues {e.code}", "detail": err_body, "url": issues_url}, 200); return

            issues_list = issues_raw.get("data", [])
            if not issues_list:
                self._json({"error": "no issues", "raw": issues_raw}, 200); return
            latest_issue = issues_list[-1] if isinstance(issues_list[-1], str) else str(issues_list[-1])

            # 2) stahni forecast
            fc_url = f"{BASE}/forecasts?issue={latest_issue}&location={location}&interval={interval}"
            try:
                raw = fetch(fc_url)
            except urllib.error.HTTPError as e:
                err_body = ""
                try: err_body = e.read().decode("utf-8")[:300]
                except: pass
                self._json({"error": f"forecasts {e.code}", "detail": err_body, "url": fc_url}, 200); return

            data = raw.get("data", [])
            points = []
            for item in data:
                dtg = item.get("dtg") or item.get("datetime")
                if dtg is None: continue
                pt = {"ts": dtg}
                for el in el_list:
                    v = item.get(el)
                    if v is not None:
                        try: pt[el] = float(v)
                        except: pt[el] = v
                points.append(pt)

            out = {
                "location": location,
                "issue": latest_issue,
                "interval": interval,
                "elements": el_list,
                "points": points,
                "n": len(points),
                "fetched_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                "_cache": "miss"
            }
            cache_all[cache_key] = {"ts": now, "data": {k: v for k, v in out.items() if k != "_cache"}}
            self._json(out)
        except Exception as e:
            print(f"  -> WEATHER ERROR: {e}", flush=True)
            self._json({"error": str(e)}, 502)

    def _forecast_de(self, qs):
        """EpexPredictor.batzill.com - statisticky model pro DE day-ahead ceny.
        DE je proxy pro CZ (Market Coupling - silna korelace).
        Vraci dalsich N hodin predikce. Cache 30 min.
        """
        try:
            if "_FORECAST_DE_CACHE" not in globals():
                globals()["_FORECAST_DE_CACHE"] = {"ts": 0, "data": None}
            cache = globals()["_FORECAST_DE_CACHE"]
            now = time.time()
            if cache["data"] and (now - cache["ts"]) < 1800:
                out = dict(cache["data"]); out["_cache"] = "hit"
                self._json(out); return

            # EpexPredictor API - region=DE, 48h forecast
            r = _request_with_retry(
                requests.get,
                "https://epexpredictor.batzill.com/prices_short",
                params={
                    "region": "DE",
                    "hours": 48,
                    "unit": "EUR_PER_MWH",
                },
                timeout=15,
                headers={"User-Agent": "Mozilla/5.0 (compatible; ceps-dashboard)"}
            )
            if r.status_code != 200:
                self._json({"error": f"EpexPredictor HTTP {r.status_code}"}, 502); return

            data = r.json()
            # Format: {"s": [unix_timestamp, ...], "t": [price, ...]}
            timestamps = data.get("s", [])
            prices = data.get("t", [])
            if not timestamps or not prices or len(timestamps) != len(prices):
                self._json({"error": "EpexPredictor returned empty/invalid data",
                            "raw_keys": list(data.keys())}, 502); return

            # Spocitej Berlin TZ aktualni hodinu pro filtrovani "next N"
            now_utc = datetime.now(timezone.utc)
            month_now = now_utc.month
            berlin_offset = 2 if 4 <= month_now <= 10 else 1
            berlin_now = now_utc + timedelta(hours=berlin_offset)
            current_unix = int(now_utc.timestamp())

            # Vyrob seznam {hour, date, price_eur} pro vsechny budouci body
            # EpexPredictor vraci unix_seconds, prevedem na Berlin TZ
            forecast_list = []
            # current_hour_start = zacatek aktualni hodiny v unix
            current_hour_start = int(berlin_now.replace(minute=0, second=0, microsecond=0).timestamp()) - berlin_offset * 3600
            for i, ts in enumerate(timestamps):
                # Vezmi hodiny od aktualni hodiny dal (vc. probihajici)
                if ts < current_hour_start - 3600:  # ignoruj jen >1h stare
                    continue
                price = prices[i]
                if price is None: continue
                point_utc = datetime.fromtimestamp(ts, tz=timezone.utc)
                point_berlin = point_utc + timedelta(hours=berlin_offset)
                forecast_list.append({
                    "unix": ts,
                    "hour": point_berlin.hour,
                    "date": point_berlin.strftime("%Y-%m-%d"),
                    "price_eur": round(float(price), 2),
                })

            # Filtr - chceme dalsich 8 hodin po aktualni (offset +1 az +8)
            # Ne probihajici hodinu (tu mas v Spot OTE KPI)
            next_8h = [f for f in forecast_list if f["unix"] > current_hour_start][:8]

            # Pad na 8 polozek pokud mame mene (napr. po publikaci nove zitra v 14:00)
            while len(next_8h) < 8:
                next_8h.append({"hour": None, "date": None, "price_eur": None, "unix": None})

            out = {
                "next_8h": next_8h,
                "next_24h": [f for f in forecast_list if f["unix"] > current_hour_start][:24],
                "all_count": len(forecast_list),
                "current_hour_start_unix": current_hour_start,
                "source": "epexpredictor.batzill.com",
                "region": "DE (proxy for CZ via Market Coupling)",
                "fetched_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                "_cache": "miss",
            }
            cache["ts"] = now
            cache["data"] = {k: v for k, v in out.items() if k != "_cache"}
            self._json(out)
        except Exception as e:
            print(f"  -> /forecast/de ERROR: {e}", flush=True)
            self._json({"error": str(e)}, 502)

    def _wind_de(self, qs):
        """Vraci aktualni vitr v severnim Nemecku - prumer ze 3 mest:
        Hamburg, Bremerhaven, Husum. Zdroj: wttr.in.
        Cache 3 hodiny + stale-while-revalidate.
        """
        try:
            if "_WIND_DE_CACHE" not in globals():
                globals()["_WIND_DE_CACHE"] = {"ts": 0, "data": None}
            cache = globals()["_WIND_DE_CACHE"]
            now = time.time()
            if cache["data"] and (now - cache["ts"]) < 10800:
                out = dict(cache["data"]); out["_cache"] = "hit"
                out["_age_sec"] = int(now - cache["ts"])
                self._json(out); return

            # Stahni pocasi pro 3 mesta v sev. Nemecku (vetrne parky)
            cities = ["Hamburg", "Bremerhaven", "Husum"]
            results = []
            for city in cities:
                d = self._fetch_wttr(city)
                if d:
                    results.append(d)

            if not results:
                # Stale fallback
                if cache["data"]:
                    out = dict(cache["data"])
                    out["_cache"] = "stale"
                    out["_age_sec"] = int(now - cache["ts"])
                    self._json(out); return
                self._json({"error": "wttr.in failed"}, 200); return

            # Helpery pro prumerovani
            def avg_or_none(values):
                vals = [v for v in values if v is not None]
                if not vals: return None
                return sum(vals) / len(vals)

            # Prumeruj napric mesty
            cur_temp = avg_or_none([d["current"]["temp_c"] for d in results])
            cur_wind = avg_or_none([d["current"]["wind_ms"] for d in results])
            cur_cloud = avg_or_none([d["current"]["cloud_pct"] for d in results])
            # Pro icon vezmeme z prvniho mesta (Hamburg)
            cur_code = results[0]["current"]["weather_code"]

            today_max = avg_or_none([d["today"]["temp_max"] for d in results])
            today_min = avg_or_none([d["today"]["temp_min"] for d in results])
            today_wind_max = avg_or_none([d["today"]["wind_max_ms"] for d in results])
            today_sun = avg_or_none([d["today"]["sunshine_h"] for d in results])
            today_icon = results[0]["today"]["icon"]
            today_desc = results[0]["today"]["desc"]

            tom_max = avg_or_none([d["tomorrow"]["temp_max"] for d in results])
            tom_min = avg_or_none([d["tomorrow"]["temp_min"] for d in results])
            tom_wind_max = avg_or_none([d["tomorrow"]["wind_max_ms"] for d in results])
            tom_sun = avg_or_none([d["tomorrow"]["sunshine_h"] for d in results])
            tom_icon = results[0]["tomorrow"]["icon"]
            tom_desc = results[0]["tomorrow"]["desc"]

            out = {
                "current": {
                    "temp_c": round(cur_temp, 1) if cur_temp is not None else None,
                    "wind_ms": round(cur_wind, 1) if cur_wind is not None else None,
                    "cloud_pct": round(cur_cloud) if cur_cloud is not None else None,
                    "weather_code": cur_code,
                },
                "today": {
                    "temp_max": round(today_max, 1) if today_max is not None else None,
                    "temp_min": round(today_min, 1) if today_min is not None else None,
                    "wind_max_ms": round(today_wind_max, 1) if today_wind_max is not None else None,
                    "icon": today_icon,
                    "desc": today_desc,
                    "sunshine_h": round(today_sun, 1) if today_sun is not None else None,
                },
                "tomorrow": {
                    "temp_max": round(tom_max, 1) if tom_max is not None else None,
                    "temp_min": round(tom_min, 1) if tom_min is not None else None,
                    "wind_max_ms": round(tom_wind_max, 1) if tom_wind_max is not None else None,
                    "icon": tom_icon,
                    "desc": tom_desc,
                    "sunshine_h": round(tom_sun, 1) if tom_sun is not None else None,
                },
                "location": "DE-sever (HH+HB+Husum)",
                "cities_ok": len(results),
                "source": "wttr.in",
                "fetched_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                "_cache": "miss",
            }
            cache["ts"] = now
            cache["data"] = {k: v for k, v in out.items() if k != "_cache"}
            self._json(out)
        except Exception as e:
            print(f"  -> wind-de ERROR: {e}", flush=True)
            self._json({"error": str(e)}, 502)

    # ============================================================
    # Spot ceny (DA) z energy-charts.info pro CZ a DE
    # ============================================================
    _SPOT_CACHE = {}  # key=(country, date), value=(ts, data)
    
    def _spot_prices(self, qs):
        """DA spot ceny z api.energy-charts.info.
        Query: ?country=CZ|DE  (default CZ)
               ?date=YYYY-MM-DD (default dnes)
        """
        try:
            country = qs.get("country", ["CZ"])[0].upper()
            date_str = qs.get("date", [datetime.now().strftime("%Y-%m-%d")])[0]
            
            bzn_map = {"CZ": "CZ", "DE": "DE-LU", "AT": "AT", "SK": "SK", "PL": "PL"}
            bzn = bzn_map.get(country, country)
            
            cache_key = (country, date_str)
            now_ts = time.time()
            cache = self._SPOT_CACHE.get(cache_key)
            if cache and (now_ts - cache[0]) < 1800:
                out = dict(cache[1])
                out["_cache"] = "hit"
                self._json(out); return
            
            url = f"https://api.energy-charts.info/price?bzn={bzn}&start={date_str}&end={date_str}"
            r = requests.get(url, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
            r.raise_for_status()
            j = r.json()
            
            unix_seconds = j.get("unix_seconds", [])
            prices = j.get("price", [])
            
            # Render bezi v UTC - musime prevest na CET/CEST (Europe/Prague)
            from zoneinfo import ZoneInfo
            tz_prg = ZoneInfo("Europe/Prague")
            
            data = []
            target_date = date_str  # YYYY-MM-DD (Prague local)
            for ts, p in zip(unix_seconds, prices):
                if p is None: continue
                dt = datetime.fromtimestamp(ts, tz=tz_prg)
                dt_date = dt.strftime("%Y-%m-%d")
                # Filtruj jen body s pozadovanym datem v Prague TZ
                if dt_date != target_date:
                    continue
                data.append({
                    "hour": dt.hour,
                    "minute": dt.minute,
                    "ts": dt.strftime("%Y-%m-%dT%H:%M"),
                    "price_eur": float(p),
                })
            
            out = {
                "country": country,
                "date": date_str,
                "data": data,
                "_cache": "miss",
            }
            self._SPOT_CACHE[cache_key] = (now_ts, out)
            print(f"  -> /spot/prices {country} {date_str}: {len(data)} bodu", flush=True)
            self._json(out)
        except Exception as e:
            print(f"  -> /spot/prices ERROR: {e}", flush=True)
            self._json({"error": str(e)}, 502)

    # ============================================================
    # OTE VDT (vnitrodenni trh CR, 15min last cena z XLSX)
    # ============================================================
    _VDT_CACHE = {}  # key=date_str, value=(ts, data)
    
    def _ote_vdt(self, qs):
        """OTE VDT z XLSX souboru.
        Query: ?date=YYYY-MM-DD (default dnes)
        Vraci: {date, data: [{ts, timestamp, last}, ...]}
        """
        try:
            date_str = qs.get("date", [datetime.now().strftime("%Y-%m-%d")])[0]
            day_dt = datetime.strptime(date_str, "%Y-%m-%d")
            nocache = qs.get("nocache", ["0"])[0] in ("1", "true", "yes")
            
            # Cache 10 min - ALE preskoc pokud prazdne (mohla byt cache z padajiciho parsingu)
            now_ts = time.time()
            cache = self._VDT_CACHE.get(date_str)
            if not nocache and cache and (now_ts - cache[0]) < 600:
                cached_data = cache[1].get("data", [])
                if cached_data:  # jen pokud cache ma data
                    out = dict(cache[1])
                    out["_cache"] = "hit"
                    self._json(out); return
            
            # 3 URL varianty - bez /view, s /view, .xls
            urls = [
                f"https://www.ote-cr.cz/attachments/27/{day_dt.year}/"
                f"month{day_dt.month:02d}/day{day_dt.day:02d}/"
                f"VDT_15MIN_{day_dt.day:02d}_{day_dt.month:02d}_{day_dt.year}_CZ.xlsx",
                f"https://www.ote-cr.cz/attachments/27/{day_dt.year}/"
                f"month{day_dt.month:02d}/day{day_dt.day:02d}/"
                f"VDT_15MIN_{day_dt.day:02d}_{day_dt.month:02d}_{day_dt.year}_CZ.xlsx/view",
            ]
            
            xlsx_bytes = None
            for url in urls:
                try:
                    r = requests.get(url, timeout=15, headers={
                        "User-Agent": "Mozilla/5.0",
                        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, */*",
                    })
                    if r.status_code == 200 and r.content[:2] == b"PK":
                        xlsx_bytes = r.content
                        print(f"  -> VDT {date_str}: XLSX nalezeno", flush=True)
                        break
                except Exception:
                    continue
            
            if xlsx_bytes is None:
                out = {"date": date_str, "data": [], "error": "no XLSX"}
                self._VDT_CACHE[date_str] = (now_ts, out)
                self._json(out); return
            
            # Parse XLSX - _xlsx_iter_rows vraci list stringu per row (pozice = sloupec)
            # Sbiram vsechny radky a hledam:
            #  - header row: obsahuje "Posledni" nebo "Last" - urci sloupec ceny
            #  - interval col: obsahuje "Interval" nebo "Cas"
            points = []
            all_rows = []
            header_row_idx = None
            last_col_idx = None
            interval_col_idx = None
            vwap_col_idx = None  # vážený průměr
            
            for row_idx, cells in enumerate(_xlsx_iter_rows(xlsx_bytes)):
                if not isinstance(cells, list):
                    cells = list(cells)
                all_rows.append(cells)
                
                if header_row_idx is None:
                    for ci, val in enumerate(cells):
                        if isinstance(val, str):
                            vlow = val.lower()
                            if "posled" in vlow or "last" in vlow:
                                last_col_idx = ci
                                header_row_idx = row_idx
                            # Vážený průměr - různé varianty
                            if "vážený" in vlow or "vazeny" in vlow or "vwap" in vlow or "weighted" in vlow:
                                vwap_col_idx = ci
                                if header_row_idx is None:
                                    header_row_idx = row_idx
                            # Někdy je to "průměrná cena"
                            if "průměrn" in vlow or "prumern" in vlow or "average" in vlow:
                                if vwap_col_idx is None:
                                    vwap_col_idx = ci
                                    if header_row_idx is None:
                                        header_row_idx = row_idx
                    if header_row_idx == row_idx:
                        for ci, val in enumerate(cells):
                            if isinstance(val, str):
                                vlow = val.lower()
                                if "interval" in vlow or "čas" in vlow or "cas" in vlow:
                                    interval_col_idx = ci
                                    break
                        # Debug log
                        print(f"  -> VDT {date_str}: header found row {row_idx}, last_col={last_col_idx}, vwap_col={vwap_col_idx}, interval_col={interval_col_idx}", flush=True)
                        print(f"     headers: {[c for c in cells if isinstance(c, str) and c.strip()]}", flush=True)
                    continue
                
                # Data row - vytahni interval a vwap (preferred) nebo last
                price_col_idx = vwap_col_idx if vwap_col_idx is not None else last_col_idx
                if price_col_idx is not None and price_col_idx < len(cells):
                    interval = None
                    if interval_col_idx is not None and interval_col_idx < len(cells):
                        interval = cells[interval_col_idx]
                    if not interval:
                        for ci in range(min(3, len(cells))):
                            v = cells[ci]
                            if isinstance(v, str) and ":" in v:
                                interval = v; break
                    
                    price_val = cells[price_col_idx]
                    
                    if isinstance(interval, str) and ":" in interval and price_val:
                        try:
                            time_part = interval.split("-")[0].strip() if "-" in interval else interval.strip()
                            parts = time_part.split(":")
                            if len(parts) < 2: continue
                            hh = int(parts[0]); mm = int(parts[1])
                            if hh > 23 or mm > 59: continue
                            ts = day_dt.replace(hour=hh, minute=mm)
                            price = float(str(price_val).replace(",", "."))
                            points.append({
                                "ts": ts.strftime("%Y-%m-%dT%H:%MZ"),
                                "timestamp": ts.strftime("%Y-%m-%dT%H:%M:%S"),
                                "last": price,
                                "price_type": "vwap" if vwap_col_idx is not None else "last"
                            })
                        except (ValueError, IndexError):
                            pass
            
            # Heuristic fallback - pokud header parser nic nenasel
            if not points and all_rows:
                # DEBUG: vypsat strukturu prvnich 10 NEPRAZDNYCH radku
                print(f"  -> VDT {date_str}: header parsing FAILED, debug rows:", flush=True)
                non_empty = [r for r in all_rows if any(c for c in r)]
                for i, row in enumerate(non_empty[:10]):
                    print(f"     row {i}: {row[:15]}", flush=True)
                
                for cells in all_rows:
                    if not cells: continue
                    # Najdi cas v jakemkoli sloupci
                    interval = None
                    time_idx = None
                    for ci, v in enumerate(cells):
                        if isinstance(v, str) and ":" in v:
                            interval = v; time_idx = ci; break
                    if not interval: continue
                    
                    # Vsechny numericke v rowu (krome casu)
                    numerics = []
                    for ci, v in enumerate(cells):
                        if ci == time_idx: continue
                        try:
                            f = float(str(v).replace(",", "."))
                            if f != 0 or not isinstance(v, str) or v.strip() in ("0", "0.0"):
                                numerics.append((ci, f))
                        except (ValueError, TypeError):
                            continue
                    if not numerics: continue
                    
                    try:
                        time_part = interval.split("-")[0].strip() if "-" in interval else interval.strip()
                        parts = time_part.split(":")
                        if len(parts) < 2: continue
                        hh = int(parts[0]); mm = int(parts[1])
                        if hh > 23 or mm > 59: continue
                        ts = day_dt.replace(hour=hh, minute=mm)
                        # Nejpravejsi numericky = "Posledni cena" typicky
                        numerics.sort(key=lambda x: x[0])
                        price = float(numerics[-1][1])
                        points.append({
                            "ts": ts.strftime("%Y-%m-%dT%H:%MZ"),
                            "timestamp": ts.strftime("%Y-%m-%dT%H:%M:%S"),
                            "last": price,
                        })
                    except (ValueError, IndexError):
                        continue
                
                if points:
                    print(f"  -> VDT {date_str}: heuristic FOUND {len(points)} points", flush=True)
            
            out = {"date": date_str, "data": points, "_cache": "miss"}
            self._VDT_CACHE[date_str] = (now_ts, out)
            print(f"  -> VDT {date_str}: {len(points)} bodu", flush=True)
            self._json(out)
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            print(f"  -> /ote/vdt ERROR: {e}\n{tb}", flush=True)
            self._json({"error": str(e), "traceback": tb.split("\n")[-3:], "data": []}, 502)

    def _ote_vdt_range(self, qs):
        """VDT data za poslednich N dni - vraci body z cache.
        Pokud nejsou v cache, vrati jen co tam je (NESTAHUJE XLSX synchronnne).
        Query: ?days_back=30
        """
        try:
            days_back = int(qs.get("days_back", ["30"])[0])
            days_back = max(1, min(60, days_back))
            
            all_points = []
            today = datetime.now()
            missing_dates = []
            
            for offset in range(days_back):
                day = today - timedelta(days=offset)
                date_str = day.strftime("%Y-%m-%d")
                cache = self._VDT_CACHE.get(date_str)
                if cache and cache[1].get("data"):
                    for p in cache[1]["data"]:
                        if "date" not in p: p["date"] = date_str
                        all_points.append(p)
                else:
                    missing_dates.append(date_str)
            
            # Setrid podle timestamp
            all_points.sort(key=lambda p: p.get("timestamp", ""))
            
            print(f"  -> VDT range {days_back}d: {len(all_points)} bodu z cache, missing: {len(missing_dates)} dni", flush=True)
            self._json({
                "days_back": days_back,
                "data": all_points,
                "missing_dates": missing_dates,
                "_note": "missing dates require /ote/vdt?date=YYYY-MM-DD to populate cache"
            })
        except Exception as e:
            import traceback
            print(f"  -> /ote/vdt/range ERROR: {e}\n{traceback.format_exc()}", flush=True)
            self._json({"error": str(e), "data": []}, 502)

    def _ceps_txt(self, qs):
        """
        Download ČEPS TXT file with REAL settled prices (not API estimate).
        Format: https://www.ceps.cz/download-data/?...
        Returns JSON kompatibilní s /api?method=OdhadovanaCenaOdchylky:
          {columns: ["Estimated price [Kč/MWh]", "Date", "Interval"], 
           rows: [{Estimated...: "4216.16", Date: "29.05.2026", Interval: "00:00-00:15"}, ...]}
        """
        import urllib.parse
        date_str = qs.get("date", [None])[0]
        if not date_str:
            self._json({"error": "missing date param"}, 400); return
        
        try:
            # Parse date YYYY-MM-DD → DD.MM.YYYY format expected by ČEPS
            from datetime import datetime as _dt
            d = _dt.strptime(date_str, "%Y-%m-%d")
            df_iso = f"{date_str}T00:00:00"
            dt_iso = f"{date_str}T23:59:59"
            
            # ČEPS TXT download URL (z reverse-engineered z ceps.cz/cs/data)
            params = {
                "method": "OdhadovanaCenaOdchylky",
                "format": "TXT",
                "version": "RT",  # real-time = reálná data
                "agregation": "QH",
                "function": "AVG",
                "dateFrom": df_iso,
                "dateTo": dt_iso,
            }
            url = "https://www.ceps.cz/download-data/?" + urllib.parse.urlencode(params)
            
            req = urllib.request.Request(url, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Accept": "text/plain,*/*;q=0.8",
                "Referer": "https://www.ceps.cz/cs/data",
            })
            with urllib.request.urlopen(req, timeout=20) as r:
                raw = r.read()
            
            # Auto-detect encoding (ČEPS používá utf-8 nebo cp1250)
            text = None
            for enc in ("utf-8", "cp1250", "iso-8859-2"):
                try:
                    text = raw.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            if text is None:
                text = raw.decode("utf-8", errors="replace")
            
            # Parse TXT formát:
            #   Verze dat;Od;Do;Agregační funkce;Agregace;
            #   reálná data;...
            #   Datum;;Odhadovaná cena [Kč/MWh];
            #   29.05.2026;00:00-00:15;4216.16;
            lines = [l.strip() for l in text.split("\n") if l.strip()]
            rows = []
            for line in lines:
                parts = [p.strip() for p in line.rstrip(";").split(";")]
                # Datový řádek má 3 části: datum, interval, cena
                if len(parts) != 3: continue
                date_part, interval, price = parts
                # Datum musí být DD.MM.YYYY
                if not (len(date_part) == 10 and date_part[2] == '.' and date_part[5] == '.'):
                    continue
                # Interval HH:MM-HH:MM
                if not (len(interval) == 11 and interval[5] == '-'):
                    continue
                rows.append({
                    "Estimated price [Kč/MWh]": price,  # string, jako z API
                    "Date": date_part,
                    "Interval": interval,
                })
            
            print(f"  -> /ceps_txt {date_str}: {len(rows)} QH", flush=True)
            self._json({
                "columns": ["Estimated price [Kč/MWh]", "Date", "Interval"],
                "rows": rows,
                "source": "ceps.cz TXT (real data)",
                "date": date_str,
            })
        except Exception as e:
            import traceback
            print(f"  -> /ceps_txt ERROR: {e}\n{traceback.format_exc()}", flush=True)
            self._json({"error": str(e), "rows": []}, 502)

    def _regelleistung_afrr_energy(self, qs):
        try:
            date_str = qs.get("date", [None])[0]
            nocache = qs.get("nocache", ["0"])[0] in ("1", "true", "yes")
            if not date_str:
                now_utc = datetime.now(timezone.utc)
                month = now_utc.month
                berlin_offset = 2 if 4 <= month <= 10 else 1
                berlin_now = now_utc + timedelta(hours=berlin_offset)
                date_str = berlin_now.strftime("%Y-%m-%d")

            try:
                datetime.strptime(date_str, "%Y-%m-%d")
            except ValueError:
                self._json({"error": f"Invalid date format: {date_str}, use YYYY-MM-DD"}, 400)
                return

            if nocache:
                key = ("aFRR", "ENERGY", date_str)
                _RL_CACHE.pop(key, None)
                print(f"  -> Regelleistung: cache invalidated for {date_str}", flush=True)

            data = get_afrr_energy_data(date_str)
            self._json(data)
        except Exception as e:
            print(f"  -> Regelleistung aFRR ENERGY ERROR: {e}", flush=True)
            import traceback
            traceback.print_exc()
            self._json({"error": str(e)}, 502)

    def _regelleistung_debug(self, qs):
        """Zjednodusene shrnuti - kolik slotu, POS/NEG bidu, sample slot keys."""
        try:
            date_str = qs.get("date", [None])[0]
            if not date_str:
                now_utc = datetime.now(timezone.utc)
                month = now_utc.month
                berlin_offset = 2 if 4 <= month <= 10 else 1
                berlin_now = now_utc + timedelta(hours=berlin_offset)
                date_str = berlin_now.strftime("%Y-%m-%d")

            data = get_afrr_energy_data(date_str)
            slots = data.get("slots", {})

            # Pocitej POS/NEG bidy a slotni klice
            slot_keys = sorted(slots.keys())
            pos_total = 0
            neg_total = 0
            slots_with_pos = 0
            slots_with_neg = 0
            slots_with_both = 0

            for k, v in slots.items():
                has_pos = bool(v.get("pos", {}).get("bids"))
                has_neg = bool(v.get("neg", {}).get("bids"))
                if has_pos:
                    slots_with_pos += 1
                    pos_total += len(v["pos"]["bids"])
                if has_neg:
                    slots_with_neg += 1
                    neg_total += len(v["neg"]["bids"])
                if has_pos and has_neg:
                    slots_with_both += 1

            # Najdi 3 sloty kde je NEG (priklad pro user)
            neg_slot_examples = []
            for k in slot_keys:
                v = slots[k]
                if v.get("neg", {}).get("bids"):
                    neg_slot_examples.append({
                        "slot": k,
                        "neg_bids_count": len(v["neg"]["bids"]),
                        "first_neg_price": v["neg"]["bids"][0]["price"],
                        "last_neg_price": v["neg"]["bids"][-1]["price"],
                    })
                    if len(neg_slot_examples) >= 3:
                        break

            # Najdi 3 sloty kde je jen POS
            pos_only_examples = []
            for k in slot_keys:
                v = slots[k]
                if v.get("pos", {}).get("bids") and not v.get("neg", {}).get("bids"):
                    pos_only_examples.append(k)
                    if len(pos_only_examples) >= 5:
                        break

            self._json({
                "date": date_str,
                "directions_available": data.get("directions_available", []),
                "total_slots": len(slot_keys),
                "first_slot": slot_keys[0] if slot_keys else None,
                "last_slot": slot_keys[-1] if slot_keys else None,
                "slots_with_pos": slots_with_pos,
                "slots_with_neg": slots_with_neg,
                "slots_with_both": slots_with_both,
                "slots_with_only_pos": slots_with_pos - slots_with_both,
                "slots_with_only_neg": slots_with_neg - slots_with_both,
                "pos_total_bids": pos_total,
                "neg_total_bids": neg_total,
                "neg_slot_examples": neg_slot_examples,
                "pos_only_slot_examples": pos_only_examples,
                "_debug_skipped": data.get("_debug", {}).get("skipped", {}),
                "_debug_sample_products": data.get("_debug", {}).get("sample_products", [])[:10],
            })
        except Exception as e:
            print(f"  -> Regelleistung DEBUG ERROR: {e}", flush=True)
            import traceback
            traceback.print_exc()
            self._json({"error": str(e)}, 502)

    def _html(self):
        html_path = os.path.join(os.path.dirname(__file__), "live_odchylky.html")
        try:
            with open(html_path, "r", encoding="utf-8") as f:
                html = f.read()
            inject = """<script>
window.addEventListener('DOMContentLoaded', () => {
  const apiUrl = window.location.origin;
  API_URL = apiUrl;
  document.getElementById('setupModal').style.display = 'none';
  loadAll();
  setRefresh(15);
});
</script>"""
            html = html.replace("</body>", inject + "\n</body>")
            body = html.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self._cors(); self.end_headers()
            self.wfile.write(body)
        except FileNotFoundError:
            self._json({"error": "live_odchylky.html not found"}, 404)

    def _json(self, data, code=200):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        try:
            self.send_response(code)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self._cors(); self.end_headers()
            self.wfile.write(body)
        except (BrokenPipeError, ConnectionResetError):
            pass

    def handle_one_request(self):
        try:
            super().handle_one_request()
        except (BrokenPipeError, ConnectionResetError):
            pass

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8765))
    public_url = os.environ.get("RENDER_EXTERNAL_URL", "").rstrip("/")

    # Warmup: stahni dnesni Regelleistung data hned po startu
    def _warmup_regelleistung():
        time.sleep(3)
        try:
            now_utc = datetime.now(timezone.utc)
            month = now_utc.month
            berlin_offset = 2 if 4 <= month <= 10 else 1
            berlin_now = now_utc + timedelta(hours=berlin_offset)
            today = berlin_now.strftime("%Y-%m-%d")
            print(f"[warmup] Pre-fetching Regelleistung aFRR ENERGY for {today}", flush=True)
            t0 = time.time()
            get_afrr_energy_data(today)
            print(f"[warmup] Done in {time.time()-t0:.1f}s", flush=True)
        except Exception as e:
            print(f"[warmup] FAIL: {e}", flush=True)
    threading.Thread(target=_warmup_regelleistung, daemon=True).start()

    if public_url:
        def keepalive():
            time.sleep(60)
            while True:
                try:
                    r = requests.get(f"{public_url}/health", timeout=10)
                    print(f"[keepalive] ping {public_url}/health -> {r.status_code}", flush=True)
                except Exception as e:
                    print(f"[keepalive] FAIL: {e}", flush=True)
                time.sleep(10 * 60)
        threading.Thread(target=keepalive, daemon=True).start()
        print(f"[keepalive] thread started, pinging {public_url}/health every 10 min", flush=True)
    else:
        print("[keepalive] RENDER_EXTERNAL_URL not set - keepalive disabled", flush=True)
    print(f"CEPS API server -> port {port} (THREADED)", flush=True)
    print(f"VERSION: regelleistung-xlsx-v32-cena-re", flush=True)
    ThreadingHTTPServer(("0.0.0.0", port), Handler).serve_forever()
